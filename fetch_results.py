#!/usr/bin/env python3
"""
Descarga resultados en vivo del Mundial 2026 y los escribe en WORLDCUP (AC/AD).
Fuente: https://worldcup26.ir/get/games (API pública, sin clave).

  python3 fetch_results.py
"""
import json
import os
import re
import sys
import unicodedata
import urllib.error
import urllib.request

import openpyxl

BASE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, BASE)

SCORERS_JSON  = os.path.join(BASE, "data", "scorers.json")
LIVE_JSON     = os.path.join(BASE, "data", "live.json")
RESULTS_JSON  = os.path.join(BASE, "data", "results.json")

from excel_sync import excel_paths as _excel_paths
from team_names import to_spanish, _norm


def _load_config():
    path = os.path.join(BASE, "update_config.json")
    defaults = {
        "fetch_live_results": True,
        "api_url": "https://worldcup26.ir/get/games",
    }
    if os.path.isfile(path):
        with open(path, encoding="utf-8") as f:
            defaults.update(json.load(f))
    return defaults


def _fetch_games(url: str, retries: int = 12, backoff: float = 3.0) -> list:
    """Fetch games from the API with retries (the server is flaky)."""
    import time as _time

    last_err = None
    for attempt in range(1, retries + 1):
        try:
            req = urllib.request.Request(url, headers={"User-Agent": "PorraLosNanos/1.0"})
            with urllib.request.urlopen(req, timeout=30) as resp:
                data = json.load(resp)
            games = data.get("games", data) if isinstance(data, dict) else data
            if not isinstance(games, list):
                raise ValueError("Respuesta API inesperada")
            if attempt > 1:
                print(f"  ✅ API respondió en intento {attempt}/{retries}")
            return games
        except Exception as e:
            last_err = e
            if attempt < retries:
                wait = backoff * attempt
                print(f"  ⚠️  Intento {attempt}/{retries} falló ({e}), reintentando en {wait:.0f}s…")
                _time.sleep(wait)
    raise last_err


def _match_key(home_es: str, away_es: str) -> str:
    return f"{home_es.strip()}-{away_es.strip()}"


def _norm_key(key: str) -> str:
    s = unicodedata.normalize("NFD", key.lower())
    return "".join(c for c in s if unicodedata.category(c) != "Mn")


def _parse_scorers(raw, team: str) -> list:
    """Parse the API scorer string into [{player, minute, own_goal, team}].

    The API uses Postgres-array-like strings, e.g.:
      '{"D. Bobadilla 7\'(OG)","F. Balogun 45\'+5\'"}'

    Patterns handled:
      "Name 7'(OG)"  → minute="7'",     own_goal=True
      "Name 7'(P)"   → minute="7'",     penalty=True
      "Name 31'"     → minute="31'",    own_goal=False
      "Name 45'+5'"  → minute="45'+5'", own_goal=False
    """
    if raw is None:
        return []
    s = str(raw).strip()
    if not s or s.lower() == "null":
        return []
    # Extract quoted segments (straight " or curly “ ” ”)
    segments = re.findall(r"[\"\u201c\u201d\u201e]([^\"\u201c\u201d\u201e]+)[\"\u201c\u201d\u201e]", s)
    if not segments:
        inner = s.strip("{}").strip()
        if inner and inner.lower() != "null":
            segments = [p.strip() for p in inner.split(",")]
    out = []
    for seg in segments:
        seg = seg.strip().strip(",").strip()
        if not seg or seg.lower() == "null":
            continue
        # Extra-time: "Name 45'+5'" or "Name 45'+ 5'" with optional (OG)/(P)
        m = re.match(r"^(.*?)\s*(\d+)\s*'?\s*\+\s*(\d+)\s*'?\s*(\(OG\)|\(P\))?\s*$", seg, re.IGNORECASE)
        if m and m.group(1).strip():
            player   = m.group(1).strip()
            minute   = f"{m.group(2)}'+{m.group(3)}'"
            tag      = (m.group(4) or "").upper()
            own_goal = tag == "(OG)"
            penalty  = tag == "(P)"
        else:
            # Regular: "Name 7'" or "Name 7'(OG)" or "Name 7'(P)"
            m = re.match(r"^(.*?)\s*(\d+)\s*'?\s*(\(OG\)|\(P\))?\s*$", seg, re.IGNORECASE)
            if m and m.group(1).strip():
                player   = m.group(1).strip()
                minute   = m.group(2) + "'"
                tag      = (m.group(3) or "").upper()
                own_goal = tag == "(OG)"
                penalty  = tag == "(P)"
            else:
                player, minute, own_goal, penalty = seg, "", False, False
        out.append({"player": player, "minute": minute, "own_goal": own_goal, "penalty": penalty, "team": team})
    return out


def _scorers_for_game(g) -> list:
    return (_parse_scorers(g.get("home_scorers"), "home")
            + _parse_scorers(g.get("away_scorers"), "away"))


def write_scorers_json(games: list) -> int:
    """Write goalscorers (with minute) keyed by 'Local-Visitante' to scorers.json."""
    scorers = {}
    for g in games:
        if str(g.get("finished", "")).upper() != "TRUE":
            continue
        home_es = to_spanish(g.get("home_team_name_en", ""))
        away_es = to_spanish(g.get("away_team_name_en", ""))
        if not home_es or not away_es:
            continue
        sc = _scorers_for_game(g)
        if sc:
            scorers[_match_key(home_es, away_es)] = sc
    os.makedirs(os.path.dirname(SCORERS_JSON), exist_ok=True)
    with open(SCORERS_JSON, "w", encoding="utf-8") as f:
        json.dump(scorers, f, ensure_ascii=False, indent=2)
    print(f"⚽ Goleadores guardados: {len(scorers)} partido(s) → {SCORERS_JSON}")
    return len(scorers)


PENALTIES_JSON = os.path.join(BASE, "data", "penalties.json")

def write_penalties_json(games: list) -> int:
    """Write penalty shootout results keyed by 'Local-Visitante' to penalties.json."""
    penalties = {}
    for g in games:
        if str(g.get("finished", "")).upper() != "TRUE":
            continue
        if "home_penalty_score" in g and "away_penalty_score" in g:
            home_es = to_spanish(g.get("home_team_name_en", ""))
            away_es = to_spanish(g.get("away_team_name_en", ""))
            if not home_es or not away_es:
                continue
            try:
                hl = int(g.get("home_penalty_score"))
                hv = int(g.get("away_penalty_score"))
                penalties[_match_key(home_es, away_es)] = {"home": hl, "away": hv}
            except (TypeError, ValueError):
                pass
    os.makedirs(os.path.dirname(PENALTIES_JSON), exist_ok=True)
    with open(PENALTIES_JSON, "w", encoding="utf-8") as f:
        json.dump(penalties, f, ensure_ascii=False, indent=2)
    print(f"🎯 Penaltis guardados: {len(penalties)} partido(s) → {PENALTIES_JSON}")
    return len(penalties)


def _is_live(g) -> bool:
    """A match is live when it has started but is not finished yet.

    The API exposes ``time_elapsed`` as "notstarted", "finished" or a live
    marker (a minute like "67'", "HT", "45'+2'", …). Live matches keep updating
    ``home_score`` / ``away_score``.
    """
    if str(g.get("finished", "")).upper() == "TRUE":
        return False
    te = str(g.get("time_elapsed", "")).strip().lower()
    if te in ("", "notstarted", "not started", "null", "none"):
        return False
    return True


def write_live_json(games: list) -> int:
    """Write currently-live scores keyed by 'Local-Visitante' to live.json.

    Rebuilt from scratch on every run, so finished/not-started matches drop out
    automatically. Official results stay in the Excel (source of truth); this
    file only feeds the provisional, in-progress overlay on the website.
    """
    live = {}
    for g in games:
        if not _is_live(g):
            continue
        home_es = to_spanish(g.get("home_team_name_en", ""))
        away_es = to_spanish(g.get("away_team_name_en", ""))
        if not home_es or not away_es:
            continue
        try:
            gl = int(g.get("home_score"))
            gv = int(g.get("away_score"))
        except (TypeError, ValueError):
            gl, gv = 0, 0
        live[_match_key(home_es, away_es)] = {
            "home":    gl,
            "away":    gv,
            "minute":  str(g.get("time_elapsed", "")).strip(),
            "scorers": _scorers_for_game(g),
        }
    os.makedirs(os.path.dirname(LIVE_JSON), exist_ok=True)
    with open(LIVE_JSON, "w", encoding="utf-8") as f:
        json.dump(live, f, ensure_ascii=False, indent=2)
    if live:
        print(f"🔴 Partidos en vivo: {len(live)} → {LIVE_JSON}")
    else:
        print(f"⚪ Sin partidos en vivo → {LIVE_JSON}")
    return len(live)


def _patch_excel_cells(path: str, sheet_name: str, updates: dict) -> list:
    """Surgically set cell values in the sheet XML.
    updates: {cell_ref: str -> value} (e.g. {"AC101": 3, "AA101": "España"})
    """
    import tempfile
    import zipfile

    with zipfile.ZipFile(path) as z:
        names = z.namelist()
        infos = z.infolist()
        contents = {n: z.read(n) for n in names}

    wbxml = contents["xl/workbook.xml"].decode("utf-8")
    rels = contents["xl/_rels/workbook.xml.rels"].decode("utf-8")
    m = (re.search(r'<sheet[^>]*name="' + re.escape(sheet_name) + r'"[^>]*?r:id="(rId\d+)"', wbxml)
         or re.search(r'<sheet[^>]*?r:id="(rId\d+)"[^>]*name="' + re.escape(sheet_name) + r'"', wbxml))
    if not m:
        raise ValueError(f"Pestaña {sheet_name} no encontrada en workbook.xml")
    rid = m.group(1)
    t = (re.search(r'Id="' + rid + r'"[^>]*Target="([^"]+)"', rels)
         or re.search(r'Target="([^"]+)"[^>]*Id="' + rid + r'"', rels))
    sp = t.group(1).lstrip("/")
    sheetpath = sp if sp.startswith("xl/") else "xl/" + sp
    xml = contents[sheetpath].decode("utf-8")

    def set_cell(xml_str, ref, value):
        pat = re.compile(r'<c r="' + re.escape(ref) + r'"([^>]*?)(/>|>.*?</c>)', re.DOTALL)

        def repl(mm):
            if isinstance(value, str):
                return f'<c r="{ref}" t="str"><v>{value}</v></c>'
            else:
                attrs = re.sub(r'\s+t="[^"]*"', "", mm.group(1))
                return f'<c r="{ref}"{attrs}><v>{value}</v></c>'

        return pat.subn(repl, xml_str, count=1)

    missing = []
    for ref, val in updates.items():
        if val is None:
            continue
        xml, n = set_cell(xml, ref, val)
        if n == 0:
            missing.append(ref)
            
    contents[sheetpath] = xml.encode("utf-8")

    fd, tmp = tempfile.mkstemp(suffix=".xlsx", dir=os.path.dirname(path) or ".")
    os.close(fd)
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for info in infos:
            zout.writestr(info, contents[info.filename])
    os.replace(tmp, path)
    return missing


def _patch_worldcup_cells(path: str, updates: dict) -> list:
    """Surgically set cell values in the WORLDCUP sheet XML.
    updates: {cell_ref: str -> value} (e.g. {"AC101": 3, "AA101": "España"})
    """
    return _patch_excel_cells(path, "WORLDCUP", updates)


def _patch_worldcup_scores(path: str, updates: dict) -> list:
    """Surgically set AC/AD score cells (and optionally names/flags) in the WORLDCUP sheet XML.
    updates: {cell_ref: str -> value}
    """
    return _patch_worldcup_cells(path, updates)


def _load_wc_match_map(wc_data) -> dict:
    """Mapa equipos (normalizado) → match_num desde WORLDCUP, sin depender de data.json."""
    match_map = {}
    if wc_data is None:
        return match_map
    for r in range(4, 148):
        home = wc_data.cell(r, 27).value
        away = wc_data.cell(r, 32).value
        mn = wc_data.cell(r, 34).value
        if not home or not away or mn is None:
            continue
        if str(home).startswith("=") or str(away).startswith("="):
            continue
        if str(home).startswith("W") or str(away).startswith("W"):
            continue
        try:
            mn_int = int(mn)
        except (TypeError, ValueError):
            continue
        key = _norm_key(_match_key(str(home).strip(), str(away).strip()))
        match_map[key] = mn_int
    return match_map


def _load_data_json_match_map() -> dict:
    import json
    import os
    match_map = {}
    try:
        data_path = os.path.join(os.path.dirname(__file__), "data.json")
        with open(data_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        for m in data.get("matches", []):
            if m.get("home") and m.get("away") and m.get("match_num"):
                k = _norm_key(_match_key(m["home"], m["away"]))
                match_map[k] = m["match_num"]
    except Exception:
        pass
    return match_map


def update_excel(games: list, path: str, label: str = "") -> int:
    """Write finished match scores into WORLDCUP columns AC (29) and AD (30)."""
    wb_ro = openpyxl.load_workbook(path, data_only=False)
    if "WORLDCUP" not in wb_ro.sheetnames:
        print(f"  ⚠ {label or path}: sin pestaña WORLDCUP, omitido")
        return 0
    wc_ro = wb_ro["WORLDCUP"]

    # Load workbook with data_only=True to get evaluated team names
    wb_data = openpyxl.load_workbook(path, data_only=True)
    wc_data = wb_data["WORLDCUP"] if "WORLDCUP" in wb_data.sheetnames else None

    # Mapa equipos → match_num desde el Excel (más fiable que data.json desactualizado)
    match_map = _load_wc_match_map(wc_data)
    if not match_map:
        match_map = _load_data_json_match_map()
    else:
        # Suplementar con data.json para partidos de eliminatorias cuyos equipos
        # aún son placeholders (W83, W84…) en el Excel y por tanto fueron ignorados
        # por _load_wc_match_map.
        for k, v in _load_data_json_match_map().items():
            if k not in match_map:
                match_map[k] = v

    row_index = {}
    row_by_match_num = {}
    if wc_data:
        for r in range(4, 148):
            # Map visual match number (AH, col 34) -> row
            mn = wc_data.cell(r, 34).value
            if mn is not None:
                try:
                    row_by_match_num[int(mn)] = r
                except (TypeError, ValueError):
                    pass

            home = wc_data.cell(r, 27).value
            away = wc_data.cell(r, 32).value
            if not home or not away:
                continue
            if str(home).startswith("=") or str(away).startswith("="):
                continue
            key = _match_key(str(home), str(away))
            row_index[_norm_key(key)] = r
    wb_data.close()

    if not row_index:
        print(f"  ⚠ Excel {label}: sin nombres de equipo en WORLDCUP "
              f"(caché de fórmulas borrada). Ábrelo y guárdalo en Excel para regenerar.")

    updates = {}
    for g in games:
        if str(g.get("finished", "")).upper() != "TRUE":
            continue
        home_es = to_spanish(g.get("home_team_name_en", ""))
        away_es = to_spanish(g.get("away_team_name_en", ""))
        try:
            gl = int(g.get("home_score", ""))
            gv = int(g.get("away_score", ""))
        except (TypeError, ValueError):
            continue

        key = _norm_key(_match_key(home_es, away_es))
        k_rev = _norm_key(_match_key(away_es, home_es))
        
        row = None
        m_num = match_map.get(key)
        if not m_num and k_rev in match_map:
            m_num = match_map.get(k_rev)
            
        if m_num:
            row = row_by_match_num.get(m_num)
            if row and match_map.get(k_rev) == m_num:
                gl, gv = gv, gl
                
        if not row:
            row = row_index.get(key)
            if not row:
                row = row_index.get(k_rev)
                if row:
                    gl, gv = gv, gl
                    
        if not row:
            print(f"  ⚠ Sin fila Excel: {home_es} vs {away_es}")
            continue

        cur_l = wc_ro.cell(row, 29).value
        cur_v = wc_ro.cell(row, 30).value
        try:
            same = cur_l is not None and cur_v is not None and int(cur_l) == gl and int(cur_v) == gv
        except (TypeError, ValueError):
            same = False
            
        ex_home = wc_ro.cell(row, 27).value
        ex_away = wc_ro.cell(row, 32).value
        needs_name_update = (row >= 101) and (str(ex_home) != home_es or str(ex_away) != away_es)

        if same and not needs_name_update:
            continue
            
        updates[f"AC{row}"] = gl
        updates[f"AD{row}"] = gv
        
        if needs_name_update:
            updates[f"AA{row}"] = home_es
            updates[f"AF{row}"] = away_es
            t1_flag = TEAM_FLAGS.get(home_es, "")
            t2_flag = TEAM_FLAGS.get(away_es, "")
            if t1_flag: updates[f"AB{row}"] = t1_flag
            if t2_flag: updates[f"AE{row}"] = t2_flag
            
        print(f"  ✓ {home_es} {gl}-{gv} {away_es}  (fila WORLDCUP {row})")

    if updates:
        missing = _patch_worldcup_scores(path, updates)
        
        # Admin names sync for Knockout matches
        admin_updates = {}
        for ref, val in updates.items():
            if ref.startswith("AA"):
                row = int(ref[2:])
                h = updates.get(f"AA{row}")
                a = updates.get(f"AF{row}")
                if h and a and row >= 101:
                    admin_row = row + 63 if row <= 116 else (row + 80 if row <= 127 else (row + 89 if row <= 135 else (row + 94 if row <= 140 else (row + 97 if row <= 143 else row + 101))))
                    admin_updates[f"K{admin_row}"] = f"{h}-{a}"
        if admin_updates:
            _patch_excel_cells(path, "ADMIN", admin_updates)

        if missing:
            print(f"  ⚠ Celdas no encontradas en XML: {', '.join(missing)}")
        
        # Number of unique rows updated
        unique_rows = len(set([k[2:] for k in updates.keys() if k.startswith("AC")]))
        print(f"💾 Excel {label} guardado ({unique_rows} partido(s) actualizado(s))")
        print(f"   → {path}")
    else:
        print(f"ℹ️  Excel {label}: sin cambios")
    return len([k for k in updates.keys() if k.startswith("AC")])


TEAM_FLAGS = {
     "Argelia": "🇩🇿", "Angola": "🇦🇴", "Argentina": "🇦🇷", "Australia": "🇦🇺",
     "Austria": "🇦🇹", "Bélgica": "🇧🇪", "Bolivia": "🇧🇴", "Brasil": "🇧🇷",
     "Camerún": "🇨🇲", "Canadá": "🇨🇦", "Chile": "🇨🇱", "China": "🇨🇳",
     "Colombia": "🇨🇴", "Costa Rica": "🇨🇷", "Croacia": "🇭🇷", "República Checa": "🇨🇿",
     "Dinamarca": "🇩🇰", "Ecuador": "🇪🇨", "Egipto": "🇪🇬", "Inglaterra": "🇬🇧",
     "Francia": "🇫🇷", "Alemania": "🇩🇪", "Ghana": "🇬🇭", "Grecia": "🇬🇷",
     "Honduras": "🇭🇳", "Hungría": "🇭🇺", "Indonesia": "🇮🇩", "Irán": "🇮🇷",
     "Irak": "🇮🇶", "Israel": "🇮🇱", "Italia": "🇮🇹", "Costa de Marfil": "🇨🇮",
     "Bosnia y Herzegovina": "🇧🇦",
     "Jamaica": "🇯🇲", "Japón": "🇯🇵", "México": "🇲🇽", "Marruecos": "🇲🇦",
     "Países Bajos": "🇳🇱", "Nueva Zelanda": "🇳🇿", "Nigeria": "🇳🇬", "Noruega": "🇳🇴",
     "Panamá": "🇵🇦", "Paraguay": "🇵🇾", "Perú": "🇵🇪", "Polonia": "🇵🇱",
     "Portugal": "🇵🇹", "Catar": "🇶🇦", "Rumanía": "🇷🇴", "Arabia Saudita": "🇸🇦",
     "Escocia": "🏴\u200d󠁧󠁢󠁳󠁣󠁴\u200d󠁿", "Senegal": "🇸🇳", "Serbia": "🇷🇸", "Eslovaquia": "🇸🇰",
     "Eslovenia": "🇸🇮", "Sudáfrica": "🇿🇦", "Corea del Sur": "🇰🇷", "España": "🇪🇸",
     "Suecia": "🇸🇪", "Suiza": "🇨🇭", "Túnez": "🇹🇳", "Turquía": "🇹🇷",
     "Ucrania": "🇺🇦", "Estados Unidos": "🇺🇸", "Uruguay": "🇺🇾", "Uzbekistán": "🇺🇿",
     "Venezuela": "🇻🇪", "Gales": "🏴\u200d󠁧󠁢󠁷󠁡\u200d󠁿", "Kenia": "🇰🇪", "Tanzania": "🇹🇿",
     "Congo RD": "🇨🇩", "RD Congo": "🇨🇩", "Emiratos Árabes Unidos": "🇦🇪"
}

def _build_bracket_match_map(results_path: str, excel_path: str) -> dict:
    """Augment the match_map with resolved future KO team pairings.

    For example, match AH=94 (W81-W82) can be resolved once we know
    the winners of matches AH=82 and AH=81 from results.json.
    The J→AH mapping is read from the Excel WORLDCUP sheet.
    """
    match_map: dict = {}

    # Load current results.json
    cache: dict = {}
    if os.path.isfile(results_path):
        try:
            with open(results_path, encoding="utf-8") as f:
                cache = json.load(f).get("by_match_num", {})
        except Exception:
            pass

    # Build J→AH mapping from Excel
    j_to_ah: dict = {}
    try:
        import openpyxl as _opx
        wb = _opx.load_workbook(excel_path, data_only=True)
        if "WORLDCUP" in wb.sheetnames:
            wc = wb["WORLDCUP"]
            for r in range(4, 148):
                j  = wc.cell(r, 10).value
                ah = wc.cell(r, 34).value
                if j is not None and ah is not None:
                    try:
                        j_to_ah[int(j)] = int(ah)
                    except (TypeError, ValueError):
                        pass
        wb.close()
    except Exception:
        pass

    ah_to_j = {v: k for k, v in j_to_ah.items()}

    # Build W-label → winner mapping from cache
    wlabel_winner: dict = {}
    for ah_str, entry in cache.items():
        try:
            ah = int(ah_str)
        except ValueError:
            continue
        winner = entry.get("winner", "")
        if winner:
            j = ah_to_j.get(ah, ah)
            wlabel_winner[f"W{j}"] = winner

    # Full WORLDCUP match map (resolved teams only)
    try:
        import openpyxl as _opx
        wb = _opx.load_workbook(excel_path, data_only=True)
        wc = wb["WORLDCUP"]
        for r in range(4, 148):
            home = wc.cell(r, 27).value
            away = wc.cell(r, 32).value
            ah   = wc.cell(r, 34).value
            if not home or not away or ah is None:
                continue
            # Resolve W-labels
            h_str = str(home).strip()
            a_str = str(away).strip()
            if h_str.startswith("W"):
                h_str = wlabel_winner.get(h_str, "")
            if a_str.startswith("W"):
                a_str = wlabel_winner.get(a_str, "")
            if not h_str or not a_str or h_str.startswith("W") or a_str.startswith("W"):
                continue
            try:
                ah_i = int(ah)
            except (TypeError, ValueError):
                continue
            key = _norm_key(_match_key(h_str, a_str))
            match_map[key] = ah_i
        wb.close()
    except Exception:
        pass

    return match_map


def write_results_json(games: list, match_map: dict) -> int:
    """Write finished match results to data/results.json (replaces Excel AC/AD writes).

    Also resolves penalty shootout winners using penalties.json.
    Returns the number of newly updated matches.
    """
    from datetime import datetime, timezone

    # Load existing results
    existing: dict = {}
    if os.path.isfile(RESULTS_JSON):
        try:
            with open(RESULTS_JSON, encoding="utf-8") as f:
                existing = json.load(f).get("by_match_num", {})
        except Exception:
            pass

    # Load penalties for KO drawn matches
    pen_data: dict = {}
    pen_path = os.path.join(BASE, "data", "penalties.json")
    if os.path.isfile(pen_path):
        try:
            with open(pen_path, encoding="utf-8") as f:
                pen_data = json.load(f)
        except Exception:
            pass

    updated = 0
    for g in games:
        if str(g.get("finished", "")).upper() != "TRUE":
            continue
        home_es = to_spanish(g.get("home_team_name_en", ""))
        away_es = to_spanish(g.get("away_team_name_en", ""))
        if not home_es or not away_es:
            continue
        try:
            gl = int(g.get("home_score", ""))
            gv = int(g.get("away_score", ""))
        except (TypeError, ValueError):
            continue

        key     = _norm_key(_match_key(home_es, away_es))
        k_rev   = _norm_key(_match_key(away_es, home_es))
        m_num   = match_map.get(key) or match_map.get(k_rev)
        if not m_num:
            print(f"  ⚠ Sin match_num para: {home_es} vs {away_es}")
            continue

        # Swap goals if the stored canonical order is reversed
        if k_rev in match_map and key not in match_map:
            gl, gv = gv, gl

        # Determine winner (includes penalties for KO draws)
        if gl > gv:
            winner = home_es
        elif gv > gl:
            winner = away_es
        else:
            pen_key     = f"{home_es}-{away_es}"
            pen_key_rev = f"{away_es}-{home_es}"
            pen = pen_data.get(pen_key) or pen_data.get(pen_key_rev)
            if pen:
                ph = int(pen.get("home", 0))
                pa = int(pen.get("away", 0))
                if pen_key in pen_data:
                    winner = home_es if ph > pa else away_es
                else:
                    winner = away_es if ph > pa else home_es
            else:
                winner = ""

        entry = {"home": home_es, "away": away_es,
                 "goals_h": gl, "goals_a": gv, "winner": winner}
        if existing.get(str(m_num)) != entry:
            existing[str(m_num)] = entry
            updated += 1
            print(f"  ✓ {home_es} {gl}-{gv} {away_es}  winner={winner or '?'}  (AH={m_num})")

    os.makedirs(os.path.dirname(RESULTS_JSON), exist_ok=True)
    payload = {
        "version": 1,
        "last_updated": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "by_match_num": existing,
    }
    with open(RESULTS_JSON, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

    if updated:
        print(f"💾 results.json actualizado ({updated} partido(s)) → {RESULTS_JSON}")
    else:
        print(f"ℹ️  results.json: sin cambios")
    return updated


def main():
    cfg = _load_config()
    if not cfg.get("fetch_live_results", True):
        print("ℹ️  fetch_live_results=false en update_config.json — omitido")
        return 0

    _, file1, file2 = _excel_paths()
    if not os.path.isfile(file1):
        print(f"❌ No encuentro: {file1}")
        return 1

    url = cfg.get("api_url", "https://worldcup26.ir/get/games")
    print(f"🌐 Descargando resultados desde {url}…")
    try:
        games = _fetch_games(url)
    except (urllib.error.URLError, TimeoutError, ValueError) as e:
        print(f"❌ Error al descargar: {e}")
        return 1

    finished = [g for g in games if str(g.get("finished", "")).upper() == "TRUE"]
    print(f"📥 {len(finished)} partido(s) finalizado(s) en la API")

    # Build match_map: data.json (primary) + bracket resolution (for future KO)
    match_map = _load_data_json_match_map()
    bracket_map = _build_bracket_match_map(RESULTS_JSON, file1)
    for k, v in bracket_map.items():
        if k not in match_map:
            match_map[k] = v
    # Also supplement with Excel-derived map for any remaining gaps
    try:
        import openpyxl as _opx
        wb_ro = _opx.load_workbook(file1, data_only=True)
        wc_data = wb_ro["WORLDCUP"] if "WORLDCUP" in wb_ro.sheetnames else None
        excel_map = _load_wc_match_map(wc_data)
        for k, v in excel_map.items():
            if k not in match_map:
                match_map[k] = v
        wb_ro.close()
    except Exception:
        pass

    # ── Write match results to results.json (no Excel writes) ──────────────
    write_results_json(games, match_map)

    # ── Side-data files (unchanged) ─────────────────────────────────────────
    write_scorers_json(games)
    write_penalties_json(games)
    write_live_json(games)

    # ── Rebuild data.json from updated results ───────────────────────────────
    try:
        import sys as _sys
        _sys.path.insert(0, BASE)
        from build_static import main as _build_main
        print("🔄 Reconstruyendo data.json…")
        _build_main()
        print("✅ data.json actualizado")
    except Exception as _e:
        print(f"⚠️  No se pudo reconstruir data.json: {_e}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
