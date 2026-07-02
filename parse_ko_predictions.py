import os, json, warnings, re
import openpyxl

warnings.filterwarnings("ignore")

BASE = os.path.dirname(os.path.abspath(__file__))
FASE_DIR = os.path.join(BASE, "data", "Fase final")
OUT_PATH = os.path.join(BASE, "data", "ko_predictions.json")
DATA_DIR = os.path.join(BASE, "data")

PRED_ROWS = (
    list(range(164, 180)) +
    list(range(200, 208)) +
    list(range(220, 224)) +
    list(range(232, 234)) +
    [244] +
    [247]
)

PLAYER_MAP = {
    "JUANCHO":  "JUANCHO",
    "LARRY":    "LARRY",
    "LUISVIR":  "LUIS/VIR",
    "MEDINA":   "MEDINA",
    "VICTOR":   "VÍCTOR",
    "CRESPO":   "CRESPO",
}


def _load_match_num_to_winner_code() -> dict:
    """Load match_num -> winner_code map from the ADMIN Excel WORLDCUP sheet.
    This uses the ADMIN (template) Excel which has W73, W76, W90 etc. as
    placeholder codes in column D — unlike player Excels which have resolved
    team names in column D.
    """
    # Locate the ADMIN Excel (may have a suffix like ' [2]')
    admin_candidates = [f for f in os.listdir(DATA_DIR) if f.startswith("ADMIN-Excel") and f.endswith(".xlsx")]
    if not admin_candidates:
        return {}
    admin_path = os.path.join(DATA_DIR, sorted(admin_candidates)[-1])
    try:
        wb_admin = openpyxl.load_workbook(admin_path, data_only=True)
        ws_admin = wb_admin["WORLDCUP"]
    except Exception:
        return {}
    result: dict = {}
    for r in range(101, 148):
        j_val = ws_admin.cell(r, 10).value   # J = match_num
        d_val = ws_admin.cell(r, 4).value    # D = winner code (e.g. 'W73', 'W90')
        if j_val is not None and d_val:
            code_str = str(d_val).strip()
            # Only accept W/L codes, not resolved team names
            if re.match(r'^[WwLl]\d+$', code_str):
                try:
                    result[int(j_val)] = code_str
                except (ValueError, TypeError):
                    pass
    return result


# Precomputed once at module import time
MATCH_NUM_TO_WINNER_CODE: dict = _load_match_num_to_winner_code()


def parse_pred_cell(raw):
    if not raw or not isinstance(raw, str):
        return None
    raw = raw.strip()
    if "\xb7" not in raw or "|" not in raw:
        return None
    try:
        name_part, rest = raw.split("\xb7", 1)
        sign, scores = rest.split("|", 1)
        if "-" not in scores:
            return None
        scores = re.sub(r"\s*\(.*?\)", "", scores).strip()
        gl, gv = scores.split("-", 1)
        return name_part.strip(), sign.strip(), int(gl.strip()), int(gv.strip())
    except Exception:
        return None


def extract_player_predictions(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Pool"]
    preds = {}
    
    # Try to extract winners from the corresponding markdown file
    md_path = xlsx_path.replace(".xlsx", ".md")
    winners = {}
    if os.path.exists(md_path):
        with open(md_path, "r", encoding="utf-8") as f:
            for line in f:
                if line.startswith("|") and "Resultado" not in line and "---" not in line:
                    parts = [p.strip() for p in line.split("|") if p.strip()]
                    if len(parts) == 4:
                        # parts: Local, Resultado, Visitante, Ganador
                        match_key = f"{parts[0]}-{parts[2]}"
                        winner_str = parts[3].replace("**", "").strip()
                        res_str = parts[1]
                        if "(" in res_str and ")" in res_str:
                            penalties = res_str[res_str.find("("):res_str.find(")")+1]
                            winner_str += f" {penalties}"
                        winners[match_key] = winner_str

    # Extract Cuadro de Honor predictions
    honor = {}
    for r in range(250, 259):
        raw = ws.cell(r, 3).value
        if raw and str(raw).strip() and str(raw).strip() != "No rellenar":
            honor[str(r)] = str(raw).strip()
    if honor:
        preds["_honor"] = honor

    POOL_TO_WORLDCUP_ROW = {
        164: 101, 165: 102, 166: 103, 167: 104, 168: 105, 169: 106, 170: 107, 171: 108,
        172: 109, 173: 110, 174: 111, 175: 112, 176: 113, 177: 114, 178: 115, 179: 116,
        200: 120, 201: 121, 202: 122, 203: 123, 204: 124, 205: 125, 206: 126, 207: 127,
        220: 131, 221: 132, 222: 133, 223: 134,
        232: 138, 233: 139,
        244: 143,
        247: 147
    }

    ROW_TO_MATCH_NUM = {
        164: 73, 165: 74, 166: 75, 167: 76, 168: 77, 169: 78, 170: 79, 171: 80,
        172: 81, 173: 82, 174: 83, 175: 84, 176: 85, 177: 86, 178: 87, 179: 88,
        200: 89, 201: 90, 202: 91, 203: 92, 204: 93, 205: 94, 206: 95, 207: 96,
        220: 97, 221: 98, 222: 99, 223: 100,
        232: 101, 233: 102,
        244: 103,
        247: 104
    }

    ws_wc = wb["WORLDCUP"]

    # ── j_to_winner: match_num -> team_name (winner of that match) ──────────
    # Built incrementally from the player's own predictions as rows are processed.
    # Starts empty — populated when R16 rows are parsed, then R8 rows, etc.
    # ── Phase row sets ──────────────────────────────────────────────────────
    R16_POOL_ROWS = set(range(164, 180))   # R16 (octavos)
    R8_POOL_ROWS  = set(range(200, 208))   # R8  (cuartos)
    R4_POOL_ROWS  = set(range(220, 224))   # R4  (semis)
    SF_POOL_ROWS  = {232, 233}             # SF  (3er/4o puesto + final)
    F_POOL_ROWS   = {244, 247}             # Final and 3rd place

    def _derive_winner(mname, sign, gl, gv, winners_dict):
        """Derive winner team name from prediction fields."""
        parts = mname.split("-", 1)
        home = parts[0].strip()
        away = parts[1].strip() if len(parts) > 1 else ""
        if sign == "1":
            return home
        elif sign == "2":
            return away
        else:  # "X" draw — try winners dict (handles penalty shootouts)
            w = winners_dict.get(mname)
            if w:
                return w.split("(")[0].strip() if "(" in w else w
            return None

    def _winner_from_row(r, ws, ROW_TO_MATCH_NUM, winners):
        """Parse row r and return (match_num, winner_team_name) or (None, None)."""
        raw = ws.cell(r, 3).value
        p = parse_pred_cell(str(raw) if raw else "")
        if not p:
            return None, None
        mn, sg, gl, gv = p
        match_num = ROW_TO_MATCH_NUM.get(r)
        if match_num is None:
            return None, None
        parts = mn.split("-", 1)
        home = parts[0].strip()
        away = parts[1].strip() if len(parts) > 1 else ""
        if sg == "1":
            w = home
        elif sg == "2":
            w = away
        else:
            w = winners.get(mn)
            if w and "(" in w:
                w = w.split("(")[0].strip()
        return match_num, w

    # Use the globally precomputed map (loaded from the ADMIN Excel which has
    # proper W-codes in column D, unlike player Excels that have team names).
    match_num_to_winner_code = MATCH_NUM_TO_WINNER_CODE

    # ── Cascaded pre-passes: build phase_winners[phase] = {match_num: team} ──
    # We need these populated BEFORE the main loop so that when we compute
    # the slot key for a phase-N match, we can reverse-look up the correct
    # phase-(N-1) match_num (and thus winner_code) for each team.
    #
    # Why cascaded? Consider cuartos (R4). The slot "W89-W90" means:
    #   W89 = winner of R8-match-90 (W73-W75 bracket)
    #   W90 = winner of R8-match-89 (W74-W77 bracket)
    # To find "which R8 match produced França?", we need R8 winners keyed by
    # the R8 match_num (89-96), NOT the R16 match_nums (73-88).
    # A shared j_to_winner with smallest-match_num preference would pick the
    # R16 entry (78→França) instead of the correct R8 entry (89→França),
    # giving winner_code W77 instead of W90.

    phase_winners: dict = {
        "R16": {},  # match_num (73-88) -> team
        "R8":  {},  # match_num (89-96) -> team
        "R4":  {},  # match_num (97-100) -> team
        "SF":  {},  # match_num (101-102) -> team
    }

    # Pre-pass R16 (rows 164-179)
    for r in sorted(R16_POOL_ROWS):
        match_num, w = _winner_from_row(r, ws, ROW_TO_MATCH_NUM, winners)
        if match_num and w:
            phase_winners["R16"][match_num] = w

    # Pre-pass R8 (rows 200-207) — needs R16 winners to resolve placeholders
    for r in sorted(R8_POOL_ROWS):
        match_num, w = _winner_from_row(r, ws, ROW_TO_MATCH_NUM, winners)
        if match_num and w:
            phase_winners["R8"][match_num] = w

    # Pre-pass R4 (rows 220-223)
    for r in sorted(R4_POOL_ROWS):
        match_num, w = _winner_from_row(r, ws, ROW_TO_MATCH_NUM, winners)
        if match_num and w:
            phase_winners["R4"][match_num] = w

    # Pre-pass SF (rows 232-233)
    for r in sorted(SF_POOL_ROWS):
        match_num, w = _winner_from_row(r, ws, ROW_TO_MATCH_NUM, winners)
        if match_num and w:
            phase_winners["SF"][match_num] = w

    def _slot_key_for_row(r, mname, phase_winners, match_num_to_winner_code,
                          wc_row, ws_wc):
        """Compute the WORLDCUP bracket slot key (e.g. 'W89-W90') for a KO row.

        For R8 rows: the home/away teams come from R16 matches, so we reverse-
        lookup in phase_winners['R16'] to get the source match_num, then convert
        to a winner_code via match_num_to_winner_code.

        For R4 rows: same idea but using phase_winners['R8'].
        For SF/Final rows: using phase_winners['R4'].
        """
        if r in R8_POOL_ROWS:
            prev = phase_winners["R16"]
        elif r in R4_POOL_ROWS:
            prev = phase_winners["R8"]
        elif r in SF_POOL_ROWS:
            prev = phase_winners["R4"]
        elif r in F_POOL_ROWS:
            prev = phase_winners["SF"]
        else:
            return None

        parts_m = mname.split("-", 1)
        h_team = parts_m[0].strip()
        a_team = parts_m[1].strip() if len(parts_m) > 1 else ""

        # Build reverse: team -> match_num from the immediately preceding phase
        rev: dict = {}
        for mn_k, mn_v in prev.items():
            if isinstance(mn_v, str):
                if mn_v not in rev or mn_k < rev[mn_v]:
                    rev[mn_v] = mn_k

        h_src = rev.get(h_team)
        a_src = rev.get(a_team)
        h_code = match_num_to_winner_code.get(h_src) if h_src is not None else None
        a_code = match_num_to_winner_code.get(a_src) if a_src is not None else None

        if h_code and a_code:
            return f"{h_code}-{a_code}"

        # Fallback: use the WORLDCUP row's placeholder pair
        if wc_row:
            ph_h = ws_wc.cell(wc_row, 1).value
            ph_a = ws_wc.cell(wc_row, 2).value
            if ph_h and ph_a:
                return f"{ph_h}-{ph_a}"
        return None

    # ── Main loop: parse all prediction rows and store with correct slot key ──
    # j_to_winner tracks running cumulative winners (used by _resolve_placeholder
    # to expand W73-style codes when players leave formulas in their cells).
    j_to_winner: dict = {**phase_winners["R16"]}   # seed with R16 so R8 can resolve

    for r in PRED_ROWS:
        raw = ws.cell(r, 3).value
        parsed = parse_pred_cell(str(raw) if raw else "")
        if not parsed:
            continue

        mname_excel, sign, gl, gv = parsed
        mname = mname_excel  # default; may be resolved below

        wc_row = POOL_TO_WORLDCUP_ROW.get(r)

        # ── For R8+ matches: resolve bracket placeholders if needed ──────────
        # Players usually write real team names (e.g. "Portugal-España") so we
        # should trust those. Only substitute teams that are still placeholders
        # like "W73" or "W84" (player left the Excel formula as-is).
        if wc_row is not None and r not in R16_POOL_ROWS:
            def _is_placeholder(name):
                return not name or bool(re.match(r'^[WwLl]\d+$', name.strip()))

            parts_excel = mname_excel.split("-", 1)
            home_excel = parts_excel[0].strip() if len(parts_excel) > 0 else ""
            away_excel = parts_excel[1].strip() if len(parts_excel) > 1 else ""

            def _resolve_placeholder(team_name):
                """If team_name is a 'W73'-style placeholder, look up winner."""
                if not _is_placeholder(team_name):
                    return team_name
                m = re.match(r'^[Ww](\d+)$', team_name.strip())
                if m:
                    code_num = int(m.group(1))
                    return j_to_winner.get(code_num)
                return None

            resolved_home = _resolve_placeholder(home_excel)
            resolved_away = _resolve_placeholder(away_excel)

            if resolved_home and resolved_away:
                mname = f"{resolved_home}-{resolved_away}"
            elif resolved_home and not _is_placeholder(away_excel):
                mname = f"{resolved_home}-{away_excel}"
            elif not _is_placeholder(home_excel) and resolved_away:
                mname = f"{home_excel}-{resolved_away}"

        winner = _derive_winner(mname, sign, gl, gv, winners)
        if winner is None and mname != mname_excel:
            fallback_winner = _derive_winner(mname_excel, sign, gl, gv, winners)
            if fallback_winner:
                corrected_teams = {t.strip() for t in mname.split("-", 1)}
                winner = fallback_winner if fallback_winner in corrected_teams else None

        pred_str = f"{mname}·{sign}|{gl}-{gv}"
        if winner:
            pred_str += f"|{winner}"

        # ── Compute slot_key using phase-specific reverse lookup ──────────────
        slot_key = None
        if wc_row and r not in R16_POOL_ROWS:
            slot_key = _slot_key_for_row(
                r, mname, phase_winners, match_num_to_winner_code, wc_row, ws_wc
            )

        # Store prediction under all relevant keys
        if wc_row and r not in R16_POOL_ROWS:
            match_num = ROW_TO_MATCH_NUM.get(r)
            if match_num is not None:
                if winner:
                    j_to_winner[int(match_num)] = winner
                preds[str(match_num)] = pred_str
            if slot_key:
                preds[slot_key] = pred_str

        elif wc_row and r in R16_POOL_ROWS:
            match_num = ROW_TO_MATCH_NUM.get(r)
            if match_num is not None:
                if winner:
                    j_to_winner[int(match_num)] = winner
                preds[str(match_num)] = pred_str

        preds[mname] = pred_str
        preds[str(r)] = pred_str  # row-number fallback

    return preds



def main():
    all_preds = {}
    for fname in sorted(os.listdir(FASE_DIR)):
        if not fname.endswith(".xlsx"):
            continue
        prefix = fname.split("_")[0].upper()
        player_name = PLAYER_MAP.get(prefix)
        if not player_name:
            print(f"  WARNING: Prefijo no reconocido: {prefix} ({fname})")
            continue
        fpath = os.path.join(FASE_DIR, fname)
        preds = extract_player_predictions(fpath)
        all_preds[player_name] = preds
        print(f"  OK {player_name}: {len(preds)} predicciones leidas")

    with open(OUT_PATH, "w", encoding="utf-8") as f:
        json.dump(all_preds, f, ensure_ascii=False, indent=2)
    print(f"\nGuardado en {OUT_PATH}")
    return all_preds


if __name__ == "__main__":
    main()
