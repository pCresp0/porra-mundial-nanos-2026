"""
Porra Mundial 'Los Nanos' 2026  –  Flask web dashboard
Reads ADMIN-Excel-Mundial_NANOS_2026 [1].xlsx (5 players) and
            ADMIN-Excel-Mundial_NANOS_2026 [2].xlsx (1 player: Crespo)
"""

import os, json, time, warnings, re
from datetime import datetime, timedelta
from flask import Flask, jsonify, send_from_directory
import openpyxl

warnings.filterwarnings("ignore")

BASE  = os.path.dirname(os.path.abspath(__file__))

def _excel_paths():
    from excel_sync import excel_paths
    return excel_paths()

ADMIN, FILE1, FILE2 = _excel_paths()

import urllib.request as _urllib_req

from fixture_data import lookup_fixture, TV_LABELS
from team_players import get_team_players, KEY_PLAYERS as _KEY_PLAYERS

# ── Club country map (club name → country flag + name) ─────────────────────────
# Used in _build_player_clubs() to enrich the scorers table.
_CLUB_COUNTRY = {
    # España
    "Real Madrid":"🇪🇸 España", "FC Barcelona":"🇪🇸 España", "Barcelona":"🇪🇸 España",
    "Atletico de Madrid":"🇪🇸 España", "Atlético de Madrid":"🇪🇸 España",
    "Athletic Club":"🇪🇸 España", "Real Sociedad":"🇪🇸 España", "Villarreal":"🇪🇸 España",
    "Sevilla":"🇪🇸 España", "Rayo Vallecano":"🇪🇸 España", "Valencia":"🇪🇸 España",
    # Inglaterra
    "Arsenal FC":"🇬🇧 Inglaterra", "Arsenal":"🇬🇧 Inglaterra",
    "Liverpool FC":"🇬🇧 Inglaterra", "Liverpool":"🇬🇧 Inglaterra",
    "Man City":"🇬🇧 Inglaterra", "Manchester City":"🇬🇧 Inglaterra",
    "Man United":"🇬🇧 Inglaterra", "Manchester United":"🇬🇧 Inglaterra",
    "Chelsea":"🇬🇧 Inglaterra", "Tottenham":"🇬🇧 Inglaterra",
    "Newcastle":"🇬🇧 Inglaterra", "Aston Villa":"🇬🇧 Inglaterra",
    "West Ham":"🇬🇧 Inglaterra", "Crystal Palace":"🇬🇧 Inglaterra",
    "Brighton":"🇬🇧 Inglaterra", "Fulham":"🇬🇧 Inglaterra",
    "Bournemouth":"🇬🇧 Inglaterra", "Leicester City":"🇬🇧 Inglaterra",
    "Nottm Forest":"🇬🇧 Inglaterra", "Ipswich Town":"🇬🇧 Inglaterra",
    # Alemania
    "Bayern München":"🇩🇪 Alemania", "FC Bayern München":"🇩🇪 Alemania",
    "Borussia Dortmund":"🇩🇪 Alemania", "RB Leipzig":"🇩🇪 Alemania",
    "Bayer Leverkusen":"🇩🇪 Alemania", "Mainz":"🇩🇪 Alemania",
    "SC Freiburg":"🇩🇪 Alemania", "Stuttgart":"🇩🇪 Alemania", "AGF Aarhus":"🇩🇰 Dinamarca",
    # Francia
    "PSG":"🇫🇷 Francia", "Paris Saint-Germain":"🇫🇷 Francia",
    "Olympique de Marsella":"🇫🇷 Francia", "OM Marseille":"🇫🇷 Francia",
    "OGC Niza":"🇫🇷 Francia", "Montpellier":"🇫🇷 Francia", "FC Nantes":"🇫🇷 Francia",
    "Monaco":"🇫🇷 Francia",
    # Italia
    "AC Milan":"🇮🇹 Italia", "FC Internazionale Milano":"🇮🇹 Italia",
    "Inter de Milán":"🇮🇹 Italia", "Juventus":"🇮🇹 Italia", "Roma":"🇮🇹 Italia",
    "Napoli":"🇮🇹 Italia", "Fiorentina":"🇮🇹 Italia", "Venezia FC":"🇮🇹 Italia",
    "Salernitana":"🇮🇹 Italia", "Aris Limassol":"🇨🇾 Chipre",
    # Portugal
    "Benfica":"🇵🇹 Portugal", "Porto":"🇵🇹 Portugal", "Sporting CP":"🇵🇹 Portugal",
    # Países Bajos
    "Feyenoord":"🇳🇱 Países Bajos", "PSV":"🇳🇱 Países Bajos", "Ajax":"🇳🇱 Países Bajos",
    "FC Utrecht":"🇳🇱 Países Bajos",
    # Bélgica
    "Beerschot":"🇧🇪 Bélgica",
    # Turquía
    "Galatasaray":"🇹🇷 Turquía", "Fenerbahçe SK":"🇹🇷 Turquía", "Fenerbahçe":"🇹🇷 Turquía",
    # Arabia / Medio Oriente
    "Al-Nassr":"🇸🇦 Arabia Saudita", "Al-Hilal":"🇸🇦 Arabia Saudita",
    "Al-Ahli":"🇸🇦 Arabia Saudita", "Al-Dawsari":"🇸🇦 Arabia Saudita",
    "Al-Talaba":"🇮🇶 Iraq", "Al-Zawraa":"🇮🇶 Iraq", "Al-Shorta":"🇮🇶 Iraq",
    "Al-Karma":"🇮🇶 Iraq", "Al-Dhafra":"🇦🇪 Emiratos", "Dibba Al-Hisn":"🇦🇪 Emiratos",
    "Al-Ahly":"🇪🇬 Egipto", "Al-Duhail":"🇶🇦 Qatar", "Al-Sadd":"🇶🇦 Qatar",
    "Al-Arabi":"🇶🇦 Qatar", "Al-Jazeera":"🇯🇴 Jordania", "Al-Faisaly":"🇯🇴 Jordania",
    "Al-Ramtha":"🇯🇴 Jordania", "Al-Najma":"🇧🇭 Bahréin",
    # Resto Europa
    "Cracovia":"🇵🇱 Polonia", "Pogoń Szczecin":"🇵🇱 Polonia",
    "Pakhtakor":"🇺🇿 Uzbekistán", "Port FC":"🇹🇭 Tailandia",
    "Viktoria Plzeň":"🇨🇿 Rep. Checa", "Persib Bandung":"🇮🇩 Indonesia",
    "Sarpsborg 08":"🇳🇴 Noruega",
    # América
    "Inter Miami":"🇺🇸 EE.UU.", "Nashville SC":"🇺🇸 EE.UU.",
    "New England Rev.":"🇺🇸 EE.UU.", "Atlanta United":"🇺🇸 EE.UU.",
    "Corinthians":"🇧🇷 Brasil", "Palmeiras":"🇧🇷 Brasil", "LDU Quito":"🇪🇨 Ecuador",
    "Club América":"🇲🇽 México", "Necaxa":"🇲🇽 México", "Club Tijuana":"🇲🇽 México",
    "Panathinaikos":"🇬🇷 Grecia",
    # África
    "Mamelodi Sundowns":"🇿🇦 Sudáfrica", "Espérance":"🇹🇳 Túnez",
    "Hafr Al-Batin":"🇸🇦 Arabia Saudita",
    # Varios / retirados
    "Al-Karma":"🇮🇶 Iraq", "AEK Larnaca":"🇨🇾 Chipre",
    "Retirado":"—", "Retirado/Amateur":"—", "Burton Albion":"🇬🇧 Inglaterra",
    "Wigan Athletic":"🇬🇧 Inglaterra", "Le Havre":"🇫🇷 Francia",
    "Hajduk Split":"🇭🇷 Croacia", "Torino":"🇮🇹 Italia",
    "FC Macarthur":"🇦🇺 Australia", "Celtic":"🏴󠁧󠁢󠁳󠁣󠁴󠁿 Escocia",
}

def _build_player_clubs():
    """Return dict {display_name: {club, club_country}} from KEY_PLAYERS."""
    import unicodedata as _ud
    out = {}
    for players in _KEY_PLAYERS.values():
        for p in players:
            full = p["name"]
            parts = full.split()
            display = f"{parts[0][0]}. {' '.join(parts[1:])}" if len(parts) >= 2 else full
            club = p.get("club", "")
            country = _CLUB_COUNTRY.get(club, "")
            # Store under both full name and display name for lookup flexibility
            entry = {"club": club, "club_country": country}
            out[full] = entry
            out[display] = entry
    return out

app = Flask(__name__)


@app.after_request
def _security_headers(response):
    """Cabeceras de seguridad básicas (no alteran el funcionamiento)."""
    response.headers.setdefault("X-Content-Type-Options", "nosniff")
    response.headers.setdefault("X-Frame-Options", "DENY")
    response.headers.setdefault("Referrer-Policy", "strict-origin-when-cross-origin")
    return response


# ── cache so we don't re-read Excel on every browser refresh ────────────────
_cache = {"data": None, "ts": 0, "error": None}
CACHE_TTL = 30  # seconds


PLAYER_COLORS = [
    "#F5C518",   # gold
    "#3B82F6",   # blue
    "#10B981",   # green
    "#F97316",   # orange
    "#A855F7",   # purple
    "#EF4444",   # red
]

PHASE_LABELS = {
    "groups":    "Fase de Grupos",
    "positions": "Posiciones Grupos",
    "q16":       "Clasificados 16avos",
    "r16":       "Dieciseisavos",
    "r8":        "Octavos",
    "r4":        "Cuartos",
    "r2":        "Semifinales",
    "r34":       "3º y 4º Puesto",
    "final":     "Final",
    "honor":     "Cuadro de Honor",
}

# ── data extraction helpers ─────────────────────────────────────────────────

def _val(ws, row, col):
    """Return cell value, converting datetime objects to ISO strings."""
    v = ws.cell(row=row, column=col).value
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    return v


def _phase_for_row(row: int) -> str:
    if 6 <= row <= 77:    return "groups"
    if 80 <= row <= 127:  return "positions"
    if 130 <= row <= 161: return "q16"
    if 163 <= row <= 179: return "r16"
    if 181 <= row <= 209: return "r8"
    if 210 <= row <= 225: return "r4"
    if 226 <= row <= 235: return "r2"
    if 236 <= row <= 238: return "r34"
    if 239 <= row <= 242: return "final"
    if 243 <= row <= 245: return "r34"
    if 246 <= row <= 248: return "final"
    return "honor"


def _parse_result(m_val):
    """Parse 'sign|score' string into dict. Returns None if not played.
    Only actual match results have the 'sign|score' pipe format."""
    if not m_val:
        return None
    s = str(m_val).strip()
    if "|" in s and s not in ("-",):
        sign, score = s.split("|", 1)
        return {"sign": sign.strip(), "score": score.strip()}
    return None  # position/team cells, TBD markers, etc.


def _parse_pred(pred_val):
    if not pred_val or str(pred_val).strip().startswith("Pegar"):
        return None
    s = str(pred_val).strip()
    # Knockout format: "EquipoLocal-EquipoVisitante·sign|score"
    # The middle dot (·) separates the team matchup from the score prediction
    pred_home = None
    pred_away = None
    if "·" in s:
        teams_part, score_part = s.split("·", 1)
        # Parse the team names (format: "TeamA-TeamB")
        if "-" in teams_part:
            # Need to handle team names with hyphens (e.g., "RD Congo", "Bosnia y Herzegovina")
            # We assume it's split at the first "-" that separates two team names
            # Use a best-effort split: try known compound names
            teams_split = teams_part.split("-", 1)
            if len(teams_split) == 2:
                pred_home = teams_split[0].strip()
                pred_away = teams_split[1].strip()
        s = score_part.strip()
    if "|" in s:
        parts = s.split("|")
        result = {"sign": parts[0].strip(), "score": parts[1].strip()}
        if len(parts) > 2:
            result["winner"] = parts[2].strip()
    else:
        result = {"sign": s, "score": s}
    if pred_home:
        result["pred_home"] = pred_home
    if pred_away:
        result["pred_away"] = pred_away
    return result


def _parse_score_parts(score_str):
    if not score_str or "-" not in str(score_str):
        return None, None
    parts = str(score_str).split("-", 1)
    try:
        return int(parts[0]), int(parts[1])
    except ValueError:
        return None, None


def _result_from_goals(gl, gv):
    """Build sign|score dict from goal counts."""
    if gl is None or gv is None:
        return None
    try:
        gl, gv = int(gl), int(gv)
    except (TypeError, ValueError):
        return None
    sign = "1" if gl > gv else ("2" if gv > gl else "X")
    return {"sign": sign, "score": f"{gl}-{gv}"}


# Transliteration aliases from the API (worldcup26.ir uses Arabic/Persian transliteration)
# Format: "api_name_lowercase_no_accents" → "canonical display name"
# Explicit overrides take priority over fuzzy matching.
_SCORER_NAME_ALIASES = {
    # Kylian Mbappé
    "kilian ambaph":  "K. Mbappé",
    "kylian mbappe":  "K. Mbappé",
    "k. mbappe":      "K. Mbappé",
    # Bradley Barcola
    "brdli barkvla":  "B. Barcola",
    "bradley barcola": "B. Barcola",
    # Add explicit overrides here for names the fuzzy cannot resolve correctly
}

# Spanish team name → FIFA code (same names used as keys in live.json / scorers.json)
_TEAM_ES_TO_FIFA = {
    "México":"MEX","Sudáfrica":"RSA","Corea del Sur":"KOR","República Checa":"CZE",
    "Canadá":"CAN","Bosnia y Herzegovina":"BIH","Catar":"QAT","Qatar":"QAT","Suiza":"SUI",
    "Brasil":"BRA","Marruecos":"MAR","Haití":"HAI","Escocia":"SCO",
    "Estados Unidos":"USA","EE.UU.":"USA","Paraguay":"PRY","Australia":"AUS","Turquía":"TUR",
    "Alemania":"GER","Curazao":"CUW","Costa de Marfil":"CIV","Ecuador":"ECU",
    "Países Bajos":"NED","Japón":"JPN","Suecia":"SWE","Túnez":"TUN",
    "Bélgica":"BEL","Egipto":"EGY","Irán":"IRN","Nueva Zelanda":"NZL",
    "España":"ESP","Cabo Verde":"CPV","Arabia Saudita":"KSA","Uruguay":"URU",
    "Francia":"FRA","Senegal":"SEN","Irak":"IRQ","Iraq":"IRQ","Noruega":"NOR",
    "Argentina":"ARG","Argelia":"DZA","Austria":"AUT","Jordania":"JOR",
    "Portugal":"POR","RD Congo":"COD","R.D. Congo":"COD","Uzbekistán":"UZB","Colombia":"COL",
    "Inglaterra":"ENG","Croacia":"HRV","Ghana":"GHA","Panamá":"PAN",
    "Dinamarca":"DEN","Grecia":"GRE","Serbia":"SRB","Nigeria":"NGA",
    "Costa Rica":"CRC","Honduras":"HON","Jamaica":"JAM","Venezuela":"VEN",
    "Chile":"CHI","Bolivia":"BOL","Perú":"PER",
}

# ── Fuzzy player lookup built lazily from team_players.KEY_PLAYERS ─────────────
_PLAYER_LOOKUP = None  # dict {fifa_code: [(norm_str, display_name), ...]}, built lazily

def _build_player_lookup():
    """Build normalised lookup from KEY_PLAYERS → dict {fifa_code: [(norm, display), ...]}"""
    import unicodedata as _ud
    try:
        from team_players import KEY_PLAYERS
    except ImportError:
        return {}
    result = {}
    for fifa_code, players in KEY_PLAYERS.items():
        entries = []
        for p in players:
            full = p["name"]
            parts = full.split()
            display = f"{parts[0][0]}. {' '.join(parts[1:])}" if len(parts) >= 2 else full
            norm = _ud.normalize("NFD", full.lower().replace("-", " "))
            norm = "".join(c for c in norm if _ud.category(c) != "Mn")
            entries.append((norm, display))
        result[fifa_code] = entries
    return result

def _fuzzy_player_match(norm_key, fifa_code=None, threshold=0.68):
    """Return display name for closest squad player, or None if below threshold.
    If fifa_code is given, only searches that team's players.
    """
    global _PLAYER_LOOKUP
    if _PLAYER_LOOKUP is None:
        _PLAYER_LOOKUP = _build_player_lookup()
    if not _PLAYER_LOOKUP:
        return None
    import difflib
    tokens = norm_key.split()
    candidates = [norm_key, " ".join(reversed(tokens))]
    # Scope to one team if known, else all teams
    if fifa_code and fifa_code in _PLAYER_LOOKUP:
        entries = _PLAYER_LOOKUP[fifa_code]
    else:
        entries = [e for team_entries in _PLAYER_LOOKUP.values() for e in team_entries]
    best_score, best_display = 0.0, None
    for norm_canon, display in entries:
        for api_str in candidates:
            s = difflib.SequenceMatcher(None, api_str, norm_canon, autojunk=False).ratio()
            if s > best_score:
                best_score, best_display = s, display
    return best_display if best_score >= threshold else None

def _fix_scorer_name(name, fifa_code=None):
    """Fix API transliteration errors in scorer names.

    1. Explicit alias dict (exact known mappings).
    2. Fuzzy match scoped to the team's squad (fifa_code) when available.
    3. Return original name if nothing matches well enough.
    """
    import unicodedata as _ud
    norm = _ud.normalize("NFD", name.strip().lower())
    key  = "".join(c for c in norm if _ud.category(c) != "Mn")
    if key in _SCORER_NAME_ALIASES:
        return _SCORER_NAME_ALIASES[key]
    fuzzy = _fuzzy_player_match(key, fifa_code=fifa_code)
    return fuzzy if fuzzy else name.strip()


def _normalize_scorer(s: dict) -> dict:
    """Normalise scorer dicts that were saved with the old broken parser.

    Old issues (caused by an incorrect regex in _parse_scorers):
    - OG baked into player: {player:"D. Bobadilla 7'(OG)", minute:""}
      → {player:"D. Bobadilla", minute:"7'", own_goal:True}
    - Extra-time split: {player:"F. Balogun 45'+", minute:"5'"}
      → {player:"F. Balogun", minute:"45'+5'", own_goal:False}
    """
    import re as _re
    player  = s.get("player", "")
    minute  = s.get("minute", "")
    own_goal = bool(s.get("own_goal", False))
    penalty  = bool(s.get("penalty", False))

    # Case 1: extra-time split — player ends with "NN'+" and minute is "M'"
    m = _re.match(r"^(.*?)\s*(\d+)\s*'\s*\+\s*$", player)
    if m:
        extra = _re.match(r"^(\d+)'?$", minute.strip())
        if extra:
            return {**s, "player": _fix_scorer_name(m.group(1).strip()),
                    "minute": f"{m.group(2)}'+{extra.group(1)}'",
                    "own_goal": own_goal, "penalty": penalty}

    # Case 2: OG embedded in player with minute: "Name 7'(OG)"
    m = _re.match(r"^(.*?)\s*(\d+)\s*'?\s*\(OG\)\s*$", player, _re.IGNORECASE)
    if m:
        return {**s, "player": _fix_scorer_name(m.group(1).strip()),
                "minute": m.group(2) + "'",
                "own_goal": True, "penalty": False}

    # Case 3: penalty embedded in player: "Name 7'(P)"
    m = _re.match(r"^(.*?)\s*(\d+)\s*'?\s*\(P\)\s*$", player, _re.IGNORECASE)
    if m:
        return {**s, "player": _fix_scorer_name(m.group(1).strip()),
                "minute": m.group(2) + "'",
                "own_goal": False, "penalty": True}

    # Already OK
    return {**s, "player": _fix_scorer_name(player), "own_goal": own_goal, "penalty": penalty}


def _load_scorers():
    """Goalscorers from data/scorers.json (written by fetch_results.py).

    Returns {normalized 'home-away' key -> [{player, minute, own_goal, team}, …]}.
    """
    path = os.path.join(BASE, "data", "scorers.json")
    out = {}
    try:
        with open(path, encoding="utf-8") as fh:
            raw = json.load(fh)
    except (FileNotFoundError, ValueError):
        return out
    for key, lst in raw.items():
        if not isinstance(lst, list):
            continue
        normalised = [_normalize_scorer(s) for s in lst]
        out[key] = normalised
        out[key.replace(" ", "")] = normalised
    return out



def _load_penalties():
    path = os.path.join(BASE, "data", "penalties.json")
    if os.path.isfile(path):
        try:
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:
            pass
    return {}



def _load_penalties():
    path = os.path.join(BASE, "data", "penalties.json")
    if os.path.isfile(path):
        try:
            with open(path, "r", encoding="utf-8") as f:
                import json
                return json.load(f)
        except Exception as e:
            pass
    return {}


def _lookup_scorers(scorers, match_name):
    if not match_name or not scorers:
        return []
    name = str(match_name).strip()
    if name in scorers:
        return scorers[name]
    compact = name.replace(" ", "")
    if compact in scorers:
        return scorers[compact]
    return []


def _load_live():
    """In-progress live scores from data/live.json (written by fetch_results.py).

    Returns {'home-away' key -> {home, away, minute, scorers}}. Only currently
    live matches are present; the file is rebuilt on every fetch run.
    """
    path = os.path.join(BASE, "data", "live.json")
    out = {}
    try:
        with open(path, encoding="utf-8") as fh:
            raw = json.load(fh)
    except (FileNotFoundError, ValueError):
        return out
    if not isinstance(raw, dict):
        return out
    for key, val in raw.items():
        if not isinstance(val, dict):
            continue
        # Extract home/away FIFA codes from the key ("Irak-Noruega" → IRQ, NOR)
        key_parts = key.split("-", 1)
        home_fifa = _TEAM_ES_TO_FIFA.get(key_parts[0].strip()) if len(key_parts) >= 1 else None
        away_fifa = _TEAM_ES_TO_FIFA.get(key_parts[1].strip()) if len(key_parts) >= 2 else None
        # Fix API transliteration errors in scorer names, scoped to the scorer's team
        scorers = val.get("scorers", [])
        if scorers and isinstance(scorers, list):
            def _fix_sc(sc):
                side = sc.get("team")  # "home" or "away"
                code = home_fifa if side == "home" else away_fifa
                return {**sc, "player": _fix_scorer_name(sc.get("player", ""), fifa_code=code)}
            scorers = [_fix_sc(sc) for sc in scorers]
            val = {**val, "scorers": scorers}
            # Recalculate score from scorers when API score lags behind
            # (API updates scorers list before updating home/away score fields)
            home_from_sc = sum(1 for sc in scorers if sc.get("team") == "home" and not sc.get("own_goal"))
            home_from_sc += sum(1 for sc in scorers if sc.get("team") == "away" and sc.get("own_goal"))
            away_from_sc = sum(1 for sc in scorers if sc.get("team") == "away" and not sc.get("own_goal"))
            away_from_sc += sum(1 for sc in scorers if sc.get("team") == "home" and sc.get("own_goal"))
            api_home = val.get("home", 0) or 0
            api_away = val.get("away", 0) or 0
            val = {**val,
                   "home": max(api_home, home_from_sc),
                   "away": max(api_away, away_from_sc)}
        out[key] = val
        out[key.replace(" ", "")] = val
    return out


def _lookup_live(live, match_name):
    if not match_name or not live:
        return None
    name = str(match_name).strip()
    if name in live:
        return live[name]
    compact = name.replace(" ", "")
    if compact in live:
        return live[compact]
    return None


RESULTS_JSON = os.path.join(BASE, "data", "results.json")


def _load_results_cache() -> dict:
    """Load match results from data/results.json (API-authoritative source).
    Returns {str(match_num): {home, away, goals_h, goals_a, winner}}
    """
    if not os.path.isfile(RESULTS_JSON):
        return {}
    try:
        with open(RESULTS_JSON, encoding="utf-8") as f:
            data = json.load(f)
        return data.get("by_match_num", {})
    except Exception:
        return {}


def _build_j_to_ah(ws_wc) -> dict:
    """Build {J_match_num → AH_match_num} mapping from WORLDCUP sheet.
    J and AH differ for 14 of the 16 R16 matches and 2 of the 8 R8 matches.
    """
    j_to_ah = {}
    for r in range(4, 148):
        j  = ws_wc.cell(r, 10).value   # J column
        ah = ws_wc.cell(r, 34).value   # AH column
        if j is not None and ah is not None:
            try:
                j_to_ah[int(j)] = int(ah)
            except (TypeError, ValueError):
                pass
    return j_to_ah


def _build_wlabel_map(results_cache: dict, j_to_ah: dict) -> dict:
    """Build {W/L-label → team name} mapping from results cache + J→AH mapping.

    W{j} = winner  of the match whose J-column value is j.
    L{j} = loser   of the match whose J-column value is j.

    We translate J→AH to look up results_cache (keyed by AH).
    """
    wl = {}
    ah_to_j = {v: k for k, v in j_to_ah.items()}
    for ah_str, entry in results_cache.items():
        try:
            ah = int(ah_str)
        except ValueError:
            continue
        j = ah_to_j.get(ah, ah)   # fall back to ah==j for matches not in map
        h = entry.get("home", "")
        a = entry.get("away", "")
        w = entry.get("winner", "")
        if not h or not a:
            continue
        # Determine loser from winner
        if w == h:
            l = a
        elif w == a:
            l = h
        else:
            # Draw with no winner recorded → can't determine loser
            l = ""
        if w:
            wl[f"W{j}"] = w
        if l:
            wl[f"L{j}"] = l
    return wl


def _build_wc_scores(filepath):
    """Goals from WORLDCUP AC/AD keyed by 'Local-Visitante' (Spanish team names).
    Falls back to Excel if results.json is absent or incomplete.
    """
    # Primary: data/results.json (API-authoritative)
    cache = _load_results_cache()
    if cache:
        scores = {}
        for _, entry in cache.items():
            h  = entry.get("home", "")
            a  = entry.get("away", "")
            gl = entry.get("goals_h")
            gv = entry.get("goals_a")
            if not h or not a or gl is None or gv is None:
                continue
            try:
                gl, gv = int(gl), int(gv)
            except (TypeError, ValueError):
                continue
            key = f"{h}-{a}"
            scores[key] = (gl, gv)
            scores[key.replace(" ", "")] = (gl, gv)
        if scores:
            return scores

    # Fallback: read from Excel WORLDCUP AC/AD columns
    wb_val = openpyxl.load_workbook(filepath, data_only=True)
    wc_val = wb_val["WORLDCUP"]
    scores = {}
    for r in range(4, 148):
        home = wc_val.cell(r, 27).value
        away = wc_val.cell(r, 32).value
        gl   = wc_val.cell(r, 29).value
        gv   = wc_val.cell(r, 30).value
        if not home or not away or str(home).startswith("=") or str(away).startswith("="):
            continue
        if gl is None or gv is None or str(gl).strip() == "" or str(gv).strip() == "":
            continue
        try:
            gl, gv = int(gl), int(gv)
        except (TypeError, ValueError):
            continue
        key = f"{str(home).strip()}-{str(away).strip()}"
        scores[key] = (gl, gv)
        scores[key.replace(" ", "")] = (gl, gv)
    wb_val.close()
    return scores


def _build_spain_times(wb):
    """Map 'Local-Visitante' → datetime España (WORLDCUP col X)."""
    wc = wb["WORLDCUP"]
    times = {}
    for r in range(4, 148):
        aa = wc.cell(row=r, column=27).value
        af = wc.cell(row=r, column=32).value
        x  = wc.cell(row=r, column=24).value
        if not aa or not af:
            continue
        if isinstance(x, datetime):
            key = f"{aa}-{af}".strip()
            times[key] = x
            times[key.replace(" ", "")] = x
    return times


def _build_wc_match_meta(wb):
    """Metadata per WORLDCUP match: teams, flags, optional scorers."""
    wc = wb["WORLDCUP"]
    meta = {}
    for r in range(4, 148):
        home = _val(wc, r, 27)  # AA
        away = _val(wc, r, 32)  # AF
        if not home or not away or str(home) in ("Casa", "Fuera", "Fecha"):
            continue
        key = f"{str(home).strip()}-{str(away).strip()}"
        fh = _val(wc, r, 28)   # AB flag
        fa = _val(wc, r, 31)   # AE flag
        # Scorer slots (if filled in Excel with player names)
        scorers = []
        for col, team in ((5, home), (6, away), (8, home), (9, away),
                          (11, home), (12, away), (14, home), (15, away)):
            v = _val(wc, r, col)
            if not v or not isinstance(v, str):
                continue
            s = v.strip()
            if s in ("-", "") or s.startswith("P.") or "Empate" in s:
                continue
            if len(s) > 2 and not s.replace(".", "").replace(" ", "").isdigit():
                scorers.append({"team": str(team).strip(), "player": s})
        ph_h = _val(wc, r, 1)  # A
        ph_a = _val(wc, r, 2)  # B
        match_num = _val(wc, r, 34)  # AH — número oficial FIFA (W-labels); col J difiere en 14 partidos
        meta[key] = {
            "home":      str(home).strip(),
            "away":      str(away).strip(),
            "home_placeholder": str(ph_h).strip() if ph_h else "",
            "away_placeholder": str(ph_a).strip() if ph_a else "",
            "flag_home": str(fh).strip() if fh else "",
            "flag_away": str(fa).strip() if fa else "",
            "scorers":   scorers,
            "match_num": int(match_num) if match_num is not None else None,
        }
        meta[key.replace(" ", "")] = meta[key]
    return meta


def _lookup_wc_meta(meta, match_name):
    if not match_name:
        return None
    name = str(match_name).strip()
    if name in meta:
        return meta[name]
    compact = name.replace(" ", "")
    if compact in meta:
        return meta[compact]
    for k, v in meta.items():
        if k.replace(" ", "") == compact:
            return v
    # fallback: split match name
    if "-" in name:
        parts = name.split("-", 1)
        return {"home": parts[0].strip(), "away": parts[1].strip(),
                "flag_home": "", "flag_away": "", "scorers": []}
    return None


def _lookup_spain_time(times, match_name):
    if not match_name:
        return None
    name = str(match_name).strip()
    if name in times:
        return times[name]
    compact = name.replace(" ", "")
    if compact in times:
        return times[compact]
    for k, v in times.items():
        if k.replace(" ", "") == compact:
            return v
    return None


def _score_breakdown(pred, result, goals_l, goals_v,
                     pts_sign=2, pts_diff=1, pts_exact=3,
                     diff_factor=1.0, multiplier=1):
    """Replica la lógica de puntuación de fase de grupos del Excel."""
    empty = {"sign": 0, "diff": 0, "exact": 0, "total": 0, "reasons": []}
    if not pred or not result:
        return empty

    pred_sign = pred.get("sign", "")
    pred_score = pred.get("score", "")
    res_sign = result.get("sign", "")
    res_score = result.get("score", "")

    if goals_l is None or goals_v is None:
        gl, gv = _parse_score_parts(res_score)
        goals_l = gl if gl is not None else 0
        goals_v = gv if gv is not None else 0

    full_pred = f"{pred_sign}|{pred_score}"
    full_res  = f"{res_sign}|{res_score}"

    if full_pred == full_res:
        reasons = [
            f"1X2 correcto (+{pts_sign})",
            f"Diferencia de goles (+{pts_diff})",
            f"Resultado exacto (+{pts_exact})",
        ]
        total = (pts_sign + pts_diff + pts_exact) * multiplier
        return {
            "sign": pts_sign * multiplier,
            "diff": pts_diff * multiplier,
            "exact": pts_exact * multiplier,
            "total": total,
            "reasons": reasons,
        }

    reasons = []
    sign_pts = diff_pts = 0

    if pred_sign == res_sign:
        sign_pts = pts_sign
        reasons.append(f"1X2 correcto (+{pts_sign})")

        pl, pv = _parse_score_parts(pred_score)
        if pl is not None and pv is not None:
            actual_diff = abs(int(goals_l) - int(goals_v))
            if pred_sign == "X":
                pred_diff = abs(pl - pv)
                diff_error = abs(actual_diff - pred_diff)
            else:
                pred_diff = abs(pl - pv)
                diff_error = abs(actual_diff - pred_diff)

            raw_diff = pts_diff * (1 - diff_error * diff_factor)
            diff_pts = max(0, round(raw_diff, 2))
            if diff_pts > 0:
                reasons.append(f"Diferencia de goles (+{diff_pts:g})")
            else:
                reasons.append("Diferencia de goles no acertada")
    else:
        reasons.append("1X2 incorrecto — 0 pts")

    total = round((sign_pts + diff_pts) * multiplier, 2)
    return {
        "sign": sign_pts * multiplier,
        "diff": diff_pts * multiplier,
        "exact": 0,
        "total": total,
        "reasons": reasons,
    }


def _week_ranges_from_dates(dates):
    """Build calendar-week filter ranges from match datetimes."""
    if not dates:
        return []
    start = min(dates).date()
    end   = max(dates).date()
    months_es = {1:"Ene",2:"Feb",3:"Mar",4:"Abr",5:"May",6:"Jun",
                 7:"Jul",8:"Ago",9:"Sep",10:"Oct",11:"Nov",12:"Dic"}
    weeks = []
    cur = start - timedelta(days=start.weekday())  # Monday
    idx = 1
    while cur <= end:
        w_end = cur + timedelta(days=6)
        m1 = months_es[cur.month]
        m2 = months_es[w_end.month]
        if cur.month == w_end.month:
            label = f"{cur.day}–{w_end.day} {m1}"
        else:
            label = f"{cur.day} {m1} – {w_end.day} {m2}"
        weeks.append({
            "id":    f"w{idx}",
            "label": label,
            "from":  cur.isoformat(),
            "to":    w_end.isoformat(),
        })
        cur += timedelta(days=7)
        idx += 1
    return weeks


STANDINGS_PHASES = [
    ("groups",    "Fase de Grupos",     "Partidos + marcadores"),
    ("positions", "Posiciones Grupos",  "1º, 2º, 3º y 4º por grupo"),
    ("q16",       "Clasificados 16avos","Equipos que pasan a dieciseisavos"),
    ("r16",       "Dieciseisavos",      "Partidos de 16avos"),
    ("r8",        "Octavos",            "Partidos de octavos"),
    ("r4",        "Cuartos",            "Partidos de cuartos"),
    ("r2",        "Semifinales",        "Partidos de semifinales"),
    ("r34_final", "3º puesto + Final",  "Partido 3º/4º y final"),
    ("honor",     "Cuadro de Honor",    "Campeón, botas, balones"),
]

_MONTHS_ES = {1:"ene",2:"feb",3:"mar",4:"abr",5:"may",6:"jun",
              7:"jul",8:"ago",9:"sep",10:"oct",11:"nov",12:"dic"}

STRENGTH_SKILLS = [
    {"key": "hits_exact",    "label": "Resultados exactos",    "icon": "🎯", "sort": "count"},
    {"key": "hits_diff",     "label": "Diferencia de goles",   "icon": "📐", "sort": "count"},
    {"key": "hits_1x2",      "label": "Signo 1X2",             "icon": "1️⃣", "sort": "count"},
    {"key": "goals_home",    "label": "Goles local exactos",   "icon": "🏠", "sort": "count"},
    {"key": "goals_away",    "label": "Goles visitante exactos","icon": "✈️", "sort": "count"},
    {"key": "avg_match",     "label": "Media pts/partido",     "icon": "⚡", "sort": "value"},
    {"key": "hit_rate",      "label": "Tasa de acierto",       "icon": "✅", "sort": "value"},
    {"key": "phase_positions","label": "Posiciones de grupos", "icon": "📊", "sort": "value"},
    {"key": "phase_q16",     "label": "Clasificados 16avos",  "icon": "🏁", "sort": "value"},
    {"key": "phase_ko",      "label": "Eliminatorias (KO)",   "icon": "⚔️", "sort": "value"},
    {"key": "phase_honor",   "label": "Cuadro de honor",      "icon": "🏆", "sort": "value"},
]

STRENGTH_BADGES = {
    "hits_exact":     "Francotirador",
    "hits_diff":      "Ojo clínico (dif.)",
    "hits_1x2":       "Rey del 1X2",
    "goals_home":     "Goles local",
    "goals_away":     "Goles visitante",
    "avg_match":      "Rendimiento puro",
    "hit_rate":       "Consistente",
    "phase_positions":"Estratega de grupos",
    "phase_q16":      "Visionario 16avos",
    "phase_ko":       "Maestro KO",
    "phase_honor":    "Oráculo del honor",
}


def _build_player_strengths(matches, standings, player_names):
    """Per-player skill stats and cross-player rankings."""
    group_played = [m for m in matches if m["phase"] == "groups" and m["played"]]
    colors = {s["name"]: s["color"] for s in standings}

    raw = {}
    for name in player_names:
        st = standings[[x["name"] for x in standings].index(name)] if name in [x["name"] for x in standings] else {}
        ko_pts = sum(float(st.get(k, 0) or 0) for k in ("r16", "r8", "r4", "r2", "r34_final"))
        raw[name] = {
            "hits_1x2": 0, "pts_1x2": 0.0,
            "hits_diff": 0, "pts_diff": 0.0,
            "hits_exact": 0, "pts_exact": 0.0,
            "goals_home": 0, "goals_away": 0,
            "matches": 0, "hits_any": 0, "pts_groups": 0.0,
            "phase_positions": float(st.get("positions", 0) or 0),
            "phase_q16": float(st.get("q16", 0) or 0),
            "phase_ko": ko_pts,
            "phase_honor": float(st.get("honor", 0) or 0),
        }

    for m in group_played:
        gl, gv = m.get("goals_l"), m.get("goals_v")
        for name in player_names:
            pd = m["predictions"].get(name, {})
            pred = pd.get("pred")
            if not pred:
                continue
            s = raw[name]
            s["matches"] += 1
            score = float(pd.get("score") or 0)
            s["pts_groups"] += score
            if score > 0:
                s["hits_any"] += 1
            brk = pd.get("breakdown") or {}
            if brk.get("sign", 0) > 0:
                s["hits_1x2"] += 1
                s["pts_1x2"] += float(brk["sign"])
            if brk.get("diff", 0) > 0:
                s["hits_diff"] += 1
                s["pts_diff"] += float(brk["diff"])
            if brk.get("exact", 0) > 0:
                s["hits_exact"] += 1
                s["pts_exact"] += float(brk["exact"])
            pl, pv = _parse_score_parts(pred.get("score", ""))
            if pl is not None and gl is not None and pl == int(gl):
                s["goals_home"] += 1
            if pv is not None and gv is not None and pv == int(gv):
                s["goals_away"] += 1

    # Build comparable values per skill
    skill_values = {sk["key"]: {} for sk in STRENGTH_SKILLS}
    for name, s in raw.items():
        n = s["matches"] or 0
        skill_values["hits_exact"][name]     = {"count": s["hits_exact"], "pts": s["pts_exact"]}
        skill_values["hits_diff"][name]      = {"count": s["hits_diff"],  "pts": s["pts_diff"]}
        skill_values["hits_1x2"][name]       = {"count": s["hits_1x2"],   "pts": s["pts_1x2"]}
        skill_values["goals_home"][name]     = {"count": s["goals_home"]}
        skill_values["goals_away"][name]     = {"count": s["goals_away"]}
        skill_values["avg_match"][name]      = {"value": round(s["pts_groups"] / n, 2) if n else 0}
        skill_values["hit_rate"][name]       = {"value": round(100 * s["hits_any"] / n) if n else 0, "hits": s["hits_any"], "total": n}
        skill_values["phase_positions"][name]= {"value": s["phase_positions"]}
        skill_values["phase_q16"][name]      = {"value": s["phase_q16"]}
        skill_values["phase_ko"][name]       = {"value": s["phase_ko"]}
        skill_values["phase_honor"][name]    = {"value": s["phase_honor"]}

    def _sort_key(skill_key, name):
        v = skill_values[skill_key][name]
        sk = next(x for x in STRENGTH_SKILLS if x["key"] == skill_key)
        if sk["sort"] == "count":
            return (v.get("count", 0), v.get("pts", 0))
        return (v.get("value", 0),)

    rankings = {}
    ranks_for_player = {name: {} for name in player_names}
    for sk in STRENGTH_SKILLS:
        key = sk["key"]
        ordered = sorted(player_names, key=lambda n: _sort_key(key, n), reverse=True)
        rows = []
        for i, name in enumerate(ordered):
            rank = i + 1
            ranks_for_player[name][key] = rank
            v = skill_values[key][name]
            if key in ("hits_exact", "hits_diff", "hits_1x2"):
                display = f"{v['count']} aciertos · {v.get('pts', 0):.0f} pts"
            elif key in ("goals_home", "goals_away"):
                display = f"{v['count']} aciertos"
            elif key == "avg_match":
                display = f"{v['value']:.2f} pts/partido"
            elif key == "hit_rate":
                display = f"{v['value']}% ({v.get('hits', 0)}/{v.get('total', 0)})"
            else:
                display = f"{v['value']:.0f} pts"
            rows.append({
                "rank": rank, "name": name, "color": colors.get(name, "#888"),
                "display": display,
                "value": v.get("count", v.get("value", 0)),
            })
        rankings[key] = rows

    players_out = []
    for st_row in standings:
        name = st_row["name"]
        s = raw[name]
        n = s["matches"]
        # Top 4 skills by rank (lowest rank number = best)
        ranked_skills = sorted(
            [(k, r) for k, r in ranks_for_player[name].items()],
            key=lambda x: (x[1], -_sort_key(x[0], name)[0] if isinstance(_sort_key(x[0], name)[0], (int, float)) else 0),
        )
        top_skills = []
        badges = []
        for key, rank in ranked_skills[:5]:
            sk = next(x for x in STRENGTH_SKILLS if x["key"] == key)
            v = skill_values[key][name]
            # Skip zero-value skills unless nothing else
            has_value = (
                v.get("count", 0) > 0 or v.get("value", 0) > 0 or v.get("pts", 0) > 0
            )
            if not has_value and len(top_skills) >= 1:
                continue
            row = next(r for r in rankings[key] if r["name"] == name)
            top_skills.append({
                "key": key, "label": sk["label"], "icon": sk["icon"],
                "rank": rank, "display": row["display"],
            })
            if rank == 1 and has_value:
                badges.append({"icon": sk["icon"], "label": STRENGTH_BADGES.get(key, sk["label"])})
            if len(top_skills) >= 4:
                break

        # Best scoring phase (from standings columns)
        phase_scores = [
            ("groups", float(st_row.get("groups", 0) or 0), "Fase de grupos"),
            ("positions", float(st_row.get("positions", 0) or 0), "Posiciones"),
            ("q16", float(st_row.get("q16", 0) or 0), "Cl. 16avos"),
            ("ko", s["phase_ko"], "Eliminatorias"),
            ("honor", s["phase_honor"], "Cuadro de honor"),
        ]
        phase_scores = [p for p in phase_scores if p[1] > 0]
        best_phase = max(phase_scores, key=lambda x: x[1]) if phase_scores else None

        players_out.append({
            "name": name,
            "color": st_row["color"],
            "pos": st_row["pos"],
            "matches_played": n,
            "stats": {
                "hits_1x2": s["hits_1x2"], "pts_1x2": s["pts_1x2"],
                "hits_diff": s["hits_diff"], "pts_diff": s["pts_diff"],
                "hits_exact": s["hits_exact"], "pts_exact": s["pts_exact"],
                "goals_home": s["goals_home"], "goals_away": s["goals_away"],
                "hit_rate": skill_values["hit_rate"][name]["value"],
                "avg_match": skill_values["avg_match"][name]["value"],
            },
            "ranks": ranks_for_player[name],
            "top_skills": top_skills,
            "badges": badges[:2],
            "best_phase": {"label": best_phase[2], "pts": best_phase[1]} if best_phase else None,
        })

    return {"players": players_out, "rankings": rankings, "skills": STRENGTH_SKILLS}


def _load_scoring_rules(ws):
    """Read all scoring criteria from ADMIN rows 8-47."""
    sections = [
        ("groups_match", "Fase de Grupos — Partidos",        range(8,  11)),
        ("groups_pos",   "Fase de Grupos — Posiciones",      range(11, 15)),
        ("q16_team",     "Clasificados Dieciseisavos",       range(15, 16)),
        ("r16",          "Dieciseisavos de Final",           range(16, 19)),
        ("r8_team",      "Clasificados Octavos",             range(19, 20)),
        ("r8",           "Octavos de Final",                 range(20, 23)),
        ("r4_team",      "Clasificados Cuartos",             range(23, 24)),
        ("r4",           "Cuartos de Final",                 range(24, 27)),
        ("r2_team",      "Clasificados Semifinales",         range(27, 28)),
        ("r2",           "Semifinales",                      range(28, 31)),
        ("r34_team",     "Clasificados 3º y 4º Puesto",      range(31, 32)),
        ("final_team",   "Clasificados Final",               range(32, 33)),
        ("r34",          "3º y 4º Puesto",                   range(33, 36)),
        ("final",        "Final",                            range(36, 39)),
        ("honor",        "Cuadro de Honor",                  range(39, 48)),
    ]
    result = []
    for key, title, rows in sections:
        items = []
        for r in rows:
            label = _val(ws, r, 3)
            pts   = _val(ws, r, 4)
            if not label or pts is None:
                continue
            try:
                pts = float(pts)
            except (ValueError, TypeError):
                continue
            # Clean label — remove redundant prefix
            lbl = str(label).strip()
            items.append({"label": lbl, "pts": pts})
        if items:
            result.append({"key": key, "title": title, "items": items})

    diff_adj = _val(ws, 50, 4)
    try:
        diff_adj = float(diff_adj) if diff_adj is not None else 0
    except (ValueError, TypeError):
        diff_adj = 0

    max_group_match = sum(i["pts"] for s in result if s["key"] == "groups_match" for i in s["items"])

    return {
        "sections":    result,
        "diff_adjustment": diff_adj,
        "max_per_group_match": max_group_match,
    }


def _parse_honor_actual(val):
    """Return real honor result or None if placeholder / TBD."""
    if val is None:
        return None
    s = str(val).strip()
    if not s or s in ("WF", "LF", "W34", "None", "-"):
        return None
    low = s.lower()
    if low.startswith("escribe") or low.startswith("pegar"):
        return None
    return _normalize_honor_name(s)


# Mapa de aliases → nombre canónico para predicciones de honor
_HONOR_NAME_ALIASES = {
    # Harry Kane
    "Kane":        "Harry Kane",
    "H. Kane":     "Harry Kane",
    "H.Kane":      "Harry Kane",
    # Cristiano Ronaldo
    "CR7":         "Cristiano Ronaldo",
    "Cristiano":   "Cristiano Ronaldo",
    "C. Ronaldo":  "Cristiano Ronaldo",
    # Vinicius Jr.
    "Vini":        "Vinicius Jr.",
    "Vini Jr":     "Vinicius Jr.",
    "Vini Jr.":    "Vinicius Jr.",
    "Vinicius":    "Vinicius Jr.",
    # Mbappé
    "Mbappe":      "Mbappé",
    "Mbappe´":     "Mbappé",
    # Oyarzabal
    "M. Oyarzabal": "Oyarzabal",
    # Haaland
    "Erling Haaland": "Haaland",
    "E. Haaland":  "Haaland",
}

def _normalize_honor_name(name):
    """Aplica aliases al nombre canónico de un jugador de honor."""
    if not name:
        return name
    return _HONOR_NAME_ALIASES.get(name.strip(), name.strip())


def _abbr_team(name):
    """3 letras en mayúscula para etiqueta compacta del eje X."""
    letters = [c for c in str(name) if c.isalpha()]
    return "".join(letters[:3]).upper() if letters else "?"


def _is_real_team_name(t):
    """True si el nombre es una selección real (no W94, 1A, 3º, etc.)."""
    if not t:
        return False
    ts = str(t).strip()
    if not ts or ts in ("3º", "4º", "ENFRENTAMIENTO FINAL"):
        return False
    if re.match(r"^[WL]\d+$", ts):
        return False
    if re.match(r"^[123][A-L]", ts):
        return False
    if "º" in ts or "finalista" in ts.lower():
        return False
    return True


def _backfill_match_flags(matches):
    """Rellena banderas faltantes en eliminatorias usando partidos anteriores."""
    flag_by_team = {}
    for m in matches:
        for side, fk in (("home", "flag_home"), ("away", "flag_away")):
            t, f = m.get(side), m.get(fk, "")
            if _is_real_team_name(t) and f:
                flag_by_team[str(t).strip()] = f
    for m in matches:
        for side, fk in (("home", "flag_home"), ("away", "flag_away")):
            t = m.get(side)
            if _is_real_team_name(t) and not m.get(fk):
                m[fk] = flag_by_team.get(str(t).strip(), "")
    return matches


def _propagate_bracket_winners(matches):
    """Sustituye Wxx/Lxx en fases siguientes por ganadores/perdedores ya conocidos."""
    by_num = {m["match_num"]: m for m in matches if m.get("match_num") is not None}

    def _winner_flag(src, winner):
        if winner == src.get("home"):
            return src.get("flag_home", "")
        if winner == src.get("away"):
            return src.get("flag_away", "")
        return ""

    def _from_w(num):
        src = by_num.get(num)
        if not src or not src.get("played") or not src.get("actual_winner"):
            return None, None
        w = str(src["actual_winner"]).strip()
        return w, _winner_flag(src, w)

    def _from_l(num):
        src = by_num.get(num)
        if not src or not src.get("played") or not src.get("actual_winner"):
            return None, None
        w = str(src["actual_winner"]).strip()
        home, away = str(src.get("home") or "").strip(), str(src.get("away") or "").strip()
        loser = away if w == home else home if w == away else ""
        if not loser:
            return None, None
        return loser, _winner_flag(src, loser)

    for m in matches:
        renamed = False
        for side, fk in (("home", "flag_home"), ("away", "flag_away")):
            ph_key = "home_placeholder" if side == "home" else "away_placeholder"
            current = str(m.get(side) or "").strip()
            # Octavos/cuartos ya fijados por nombre real: no pisar con W86/W88 del cuadro.
            if _is_real_team_name(current):
                continue
            slots = []
            if re.match(r"^[WL]\d+$", current):
                slots.append(current)
            ph = str(m.get(ph_key) or "").strip()
            if ph and ph not in slots:
                slots.append(ph)
            for slot in slots:
                wm = re.match(r"^W(\d+)$", slot)
                lm = re.match(r"^L(\d+)$", slot)
                resolved = None
                if wm:
                    resolved = _from_w(int(wm.group(1)))
                elif lm:
                    resolved = _from_l(int(lm.group(1)))
                if resolved and resolved[0]:
                    m[side] = resolved[0]
                    if resolved[1]:
                        m[fk] = resolved[1]
                    renamed = True
                    break

        h, a = str(m.get("home") or "").strip(), str(m.get("away") or "").strip()
        if renamed and h and a:
            m["name"] = f"{h}-{a}"

    _backfill_match_flags(matches)
    return matches


def _prog_match_earned(m, pred_obj):
    """Puntos de un partido para la progresión (misma lógica que la tabla de clasificación)."""
    phase = m.get("phase", "groups")
    if phase == "groups":
        return round(float(pred_obj.get("score") or 0), 2)
    bd = pred_obj.get("breakdown") or {}
    match_pts = round((bd.get("sign") or 0) + (bd.get("diff") or 0) + (bd.get("exact") or 0), 2)
    qual_pts = round(float(pred_obj.get("qual_pts") or 0), 2)
    if match_pts > 0:
        return round(match_pts + qual_pts, 2)
    score = round(float(pred_obj.get("score") or 0), 2)
    if score > 0:
        return score
    return qual_pts


def _sync_prog_players(progression, player_names):
    """Recalcula la serie acumulada de cada jugador a partir de day_points."""
    players_out = {n: [] for n in player_names}
    for name in player_names:
        cumulative = 0.0
        for pts in progression.get("day_points", {}).get(name, []):
            cumulative = round(cumulative + float(pts or 0), 1)
            players_out[name].append(cumulative)
    progression["players"] = players_out
    return progression


def _normalize_player_grid_preds(grid_preds):
    """Excel usa VICTOR; la web usa VÍCTOR."""
    if "VICTOR" in grid_preds and "VÍCTOR" not in grid_preds:
        grid_preds["VÍCTOR"] = grid_preds.pop("VICTOR")
    return grid_preds


def _progression_grid_context_from_data(dj):
    """Construye grid_ctx para progresión a partir de Excel + partidos jugados."""
    global FILE1, FILE2
    FILE1 = os.path.join(BASE, "data", "Admin real", "ALL_ADMIN-Excel-Mundial-2026_1.xlsx")
    FILE2 = os.path.join(BASE, "data", "Admin real", "ALL_ADMIN-Excel-Mundial-2026_2.xlsx")
    if not os.path.isfile(FILE1) or not os.path.isfile(FILE2):
        return None
    ws1, players1, _ = _load_file1()
    ws2, player2, _ = _load_file2()
    all_players = players1 + [player2]
    all_ws = [ws1] * len(players1) + [ws2]
    grid_preds = _normalize_player_grid_preds(_build_player_grid_preds(all_players, all_ws))
    actual_r8, actual_r4, actual_r2, actual_r34, actual_final = (set(), set(), set(), set(), set())
    _augment_actual_qualifiers_from_matches(
        dj.get("matches", []), actual_r8, actual_r4, actual_r2, actual_r34, actual_final,
    )
    return {
        "grid_preds": grid_preds,
        "pts": {
            "r8": float(_val(ws1, 19, 4) or 3.0),
            "r4": float(_val(ws1, 23, 4) or 5.0),
            "r2": float(_val(ws1, 27, 4) or 8.0),
            "r34": float(_val(ws1, 31, 4) or 10.0),
            "final": float(_val(ws1, 35, 4) or 15.0),
        },
        "actual": {
            "r8": actual_r8, "r4": actual_r4, "r2": actual_r2,
            "r34": actual_r34, "final": actual_final,
        },
    }


def _ko_match_qual_bonus(m, player_names, qual_pts):
    """Bonus «clasificado correcto» del partido (sin duplicar el cuadro de equipos)."""
    if m.get("phase") not in ("r16", "r8", "r4", "r2", "r34", "final"):
        return {}
    winner = str(m.get("actual_winner") or "").strip()
    if not winner:
        return {}
    earned = {}
    for n in player_names:
        pred_obj = m.get("predictions", {}).get(n, {})
        if float(pred_obj.get("qual_pts") or 0) > 0:
            earned[n] = float(pred_obj["qual_pts"])
            continue
        pred = pred_obj.get("pred") or {}
        if (pred_obj.get("breakdown") or {}).get("team_match") == "none":
            continue
        pred_w = str(pred.get("winner") or "").strip()
        if pred_w and pred_w.lower() == winner.lower():
            earned[n] = qual_pts
    return earned


def _append_progression_event(prog, player_names, label, flag, date, title, phase, earned):
    """Añade un evento al final de la progresión existente."""
    prog.setdefault("labels", []).append(label)
    prog.setdefault("flag_labels", []).append(flag)
    prog.setdefault("dates", []).append(date)
    prog.setdefault("titles", []).append(title)
    prog.setdefault("phases", []).append(phase)
    for n in player_names:
        prog.setdefault("day_points", {}).setdefault(n, []).append(
            round(float(earned.get(n, 0.0)), 1)
        )
    return _sync_prog_players(prog, player_names)


def _insert_progression_events(prog, idx, events, player_names):
    """Sustituye el evento en idx por una lista de eventos (misma suma de puntos)."""
    for key in ("labels", "flag_labels", "dates", "titles", "phases"):
        lst = prog.get(key, [])
        if idx < len(lst):
            lst.pop(idx)
        for offset, ev in enumerate(events):
            lst.insert(idx + offset, ev[key])
    for n in player_names:
        dp = prog.setdefault("day_points", {}).setdefault(n, [])
        if idx < len(dp):
            dp.pop(idx)
        for offset, ev in enumerate(events):
            dp.insert(idx + offset, round(float(ev.get("earned", {}).get(n, 0.0)), 1))
    return _sync_prog_players(prog, player_names)


def _match_prog_title(m):
    gl, gv = m.get("goals_l"), m.get("goals_v")
    score = f" {gl}-{gv} " if gl is not None and gv is not None else " vs "
    title = f"{m.get('home', '')}{score}{m.get('away', '')}"
    if m.get("date"):
        try:
            dt = datetime.strptime(m["date"], "%Y-%m-%d")
            title += f" · {dt.day} {_MONTHS_ES[dt.month]}"
        except ValueError:
            pass
    return title


def repair_progression(dj):
    """Corrige progresión: sincroniza players, fechas y eventos agrupados de octavos."""
    prog = dj.get("progression")
    if not prog:
        return dj
    player_names = dj.get("meta", {}).get("players", [])
    if not player_names:
        return dj

    # Fechas inválidas → fecha del partido o última válida
    dates = prog.get("dates", [])
    last_ok = next((d for d in reversed(dates) if str(d).startswith("2026-")), "2026-07-07")
    for i, d in enumerate(dates):
        if not str(d).startswith("2026-"):
            m = None
            lbl = (prog.get("labels") or [""])[i] if i < len(prog.get("labels", [])) else ""
            for cand in dj.get("matches", []):
                ab = f"{_abbr_team(cand.get('home'))}-{_abbr_team(cand.get('away'))}"
                if ab == lbl or cand.get("name", "").startswith(lbl[:3]):
                    m = cand
                    break
            dates[i] = (m.get("date") if m and str(m.get("date", "")).startswith("2026-")
                        else last_ok)

    # Desagrupar evento EST-BÉL que incluye varios partidos del 07/07
    labels = prog.get("labels", [])
    bundled_idx = next(
        (i for i, lbl in enumerate(labels)
         if lbl in ("EST-BÉL", "EST-BEL") and i < len(dates)
         and not str(dates[i]).startswith("2026-")),
        None,
    )
    if bundled_idx is None:
        bundled_idx = next(
            (i for i, lbl in enumerate(labels) if lbl in ("EST-BÉL", "EST-BEL")), None
        )

    grid_ctx = _progression_grid_context_from_data(dj)
    if bundled_idx is not None and grid_ctx:
        matches_by_name = {m.get("name"): m for m in dj.get("matches", [])}
        m_est = matches_by_name.get("Estados Unidos-Bélgica")
        m_arg = matches_by_name.get("Argentina-Egipto")
        already_split = (
            bundled_idx + 1 < len(labels)
            and labels[bundled_idx + 1] in ("ARG-EGI", "ARG-EGY")
        )
        if (m_est and m_arg and m_est.get("played") and m_arg.get("played")
                and not already_split):
            players = player_names
            grid_preds = grid_ctx["grid_preds"]
            pts_cfg = grid_ctx["pts"]
            actual = grid_ctx["actual"]
            awarded = set()
            date = m_est.get("date") or m_arg.get("date") or "2026-07-07"

            est_earned = {
                n: round(float(m_est["predictions"].get(n, {}).get("score") or 0)
                         + float(m_est["predictions"].get(n, {}).get("qual_pts") or 0), 1)
                for n in players
            }
            bel_grid = _award_grid_on_ko_win(
                m_est, players, grid_preds, pts_cfg, awarded, actual,
            )
            arg_earned = {
                n: round(float(m_arg["predictions"].get(n, {}).get("score") or 0), 1)
                for n in players
            }
            arg_grid = _award_grid_on_ko_win(
                m_arg, players, grid_preds, pts_cfg, awarded, actual,
            )
            arg_qual = _ko_match_qual_bonus(m_arg, players, pts_cfg["r4"])
            arg_combined = {
                n: round(arg_grid.get(n, 0) + arg_qual.get(n, 0), 1) for n in players
            }

            base_totals = {
                n: round(
                    est_earned.get(n, 0) + bel_grid.get(n, 0)
                    + arg_earned.get(n, 0) + arg_combined.get(n, 0),
                    1,
                )
                for n in players
            }
            old_totals = {n: prog["day_points"][n][bundled_idx] for n in players}
            if not all(abs(old_totals[n] - base_totals[n]) < 0.05 for n in players):
                # Ajustar bonus de clasificado para respetar totales ya consolidados
                arg_combined = {
                    n: round(
                        old_totals[n] - est_earned.get(n, 0) - bel_grid.get(n, 0)
                        - arg_earned.get(n, 0),
                        1,
                    )
                    for n in players
                }

            events = [
                {
                    "labels": f"{_abbr_team(m_est.get('home'))}-{_abbr_team(m_est.get('away'))}",
                    "flag_labels": f"{m_est.get('flag_home', '')}{m_est.get('flag_away', '')}",
                    "dates": date, "titles": _match_prog_title(m_est),
                    "phases": "r8", "earned": est_earned,
                },
            ]
            if any(v > 0 for v in bel_grid.values()):
                w = str(m_est.get("actual_winner") or "").strip()
                events.append({
                    "labels": f"Clasif. {_abbr_team(w)}",
                    "flag_labels": "✓", "dates": date,
                    "titles": f"Clasificado a fase siguiente: {w}",
                    "phases": "r4_team", "earned": bel_grid,
                })
            events.append({
                "labels": f"{_abbr_team(m_arg.get('home'))}-{_abbr_team(m_arg.get('away'))}",
                "flag_labels": f"{m_arg.get('flag_home', '')}{m_arg.get('flag_away', '')}",
                "dates": date, "titles": _match_prog_title(m_arg),
                "phases": "r8", "earned": arg_earned,
            })
            if any(v > 0 for v in arg_combined.values()):
                w = str(m_arg.get("actual_winner") or "").strip()
                events.append({
                    "labels": f"Clasif. {_abbr_team(w)}",
                    "flag_labels": "✓", "dates": date,
                    "titles": f"Clasificado a fase siguiente: {w}",
                    "phases": "r4_team", "earned": arg_combined,
                })

            old_totals = {n: prog["day_points"][n][bundled_idx] for n in players}
            new_totals = {
                n: round(
                    est_earned.get(n, 0) + bel_grid.get(n, 0)
                    + arg_earned.get(n, 0) + arg_combined.get(n, 0),
                    1,
                )
                for n in players
            }
            if all(abs(old_totals[n] - new_totals[n]) < 0.05 for n in players):
                _insert_progression_events(prog, bundled_idx, events, players)
            else:
                _sync_prog_players(prog, players)
        else:
            _sync_prog_players(prog, players)
    else:
        _sync_prog_players(prog, players)

    dj["progression"] = prog
    return dj


def _ko_match_qual_target_phase(m):
    """Columna de standings donde van los pts de clasificado del partido."""
    return {"r16": "r8", "r8": "r4", "r4": "r2", "r2": "r34_final",
            "r34": "r34_final", "final": "r34_final"}.get(m.get("phase"))


def _augment_actual_qualifiers_from_matches(matches, actual_r8, actual_r4, actual_r2,
                                            actual_r34, actual_final):
    """Añade ganadores ya conocidos cuando el Excel aún muestra placeholders Wxx."""
    for m in matches:
        if not m.get("played"):
            continue
        w = m.get("actual_winner")
        if not w:
            continue
        w = str(w).strip()
        ph = m.get("phase")
        if ph == "r16":
            actual_r8.add(w)
        elif ph == "r8":
            actual_r4.add(w)
        elif ph == "r4":
            actual_r2.add(w)
        elif ph == "r2":
            actual_final.add(w)
            home = str(m.get("home") or "").strip()
            away = str(m.get("away") or "").strip()
            loser = away if w == home else home if w == away else ""
            if loser:
                actual_r34.add(loser)
        elif ph == "r34":
            actual_r34.add(w)
        elif ph == "final":
            actual_final.add(w)


def _build_player_grid_preds(all_players, all_ws):
    """Equipos predichos en cada cuadro de clasificados (filas ADMIN)."""
    ranges = {
        "r8":    range(182, 198),
        "r4":    range(210, 218),
        "r2":    range(226, 230),
        "r34":   range(236, 238),
        "final": range(240, 242),
    }
    out = {}
    for p, ws in zip(all_players, all_ws):
        name = p["name"]
        pc = p["pred_col"]
        out[name] = {}
        for key, rows in ranges.items():
            out[name][key] = {
                str(_val(ws, r, pc) or "").strip()
                for r in rows if _val(ws, r, pc)
            }
    return out


def _append_prog_step(cumulative, players_out, day_points, labels, flag_labels,
                      dates, titles, phases, player_names, label, flag, date,
                      title, phase, earned_by_player):
    """Añade un paso (partido o hito) a la serie de progresión."""
    labels.append(label)
    flag_labels.append(flag)
    dates.append(date)
    titles.append(title)
    phases.append(phase)
    for n in player_names:
        e = round(float(earned_by_player.get(n, 0.0)), 1)
        cumulative[n] = round(cumulative[n] + e, 1)
        players_out[n].append(cumulative[n])
        day_points[n].append(e)


def _award_grid_on_ko_win(m, player_names, grid_preds, pts_cfg, awarded_grid, actual_sets=None):
    """Puntos de cuadro de clasificados al clasificar un equipo (misma lógica que CLAS)."""
    phase = m.get("phase")
    winner = m.get("actual_winner")
    if not winner:
        return {}
    actual_sets = actual_sets or {}
    winner = str(winner).strip()
    earned = {n: 0.0 for n in player_names}

    if phase == "r16":
        grid_key, pt_key = "r8", "r8"
    elif phase == "r8":
        grid_key, pt_key = "r4", "r4"
    elif phase == "r4":
        grid_key, pt_key = "r2", "r2"
    elif phase == "r2":
        home = str(m.get("home") or "").strip()
        away = str(m.get("away") or "").strip()
        loser = away if winner == home else home if winner == away else ""
        actual_final = actual_sets.get("final", set())
        actual_r34 = actual_sets.get("r34", set())
        pt_val_f = pts_cfg.get("final", 0)
        pt_val_34 = pts_cfg.get("r34", 0)
        for n in player_names:
            preds = grid_preds.get(n, {})
            if winner and winner in preds.get("final", set()) and winner in actual_final:
                key = (n, winner, "final")
                if key not in awarded_grid:
                    awarded_grid.add(key)
                    earned[n] += pt_val_f
            if loser and loser in preds.get("r34", set()) and loser in actual_r34:
                key = (n, loser, "r34")
                if key not in awarded_grid:
                    awarded_grid.add(key)
                    earned[n] += pt_val_34
        return earned
    else:
        return earned

    actual = actual_sets.get(grid_key, set())
    pt_val = pts_cfg.get(pt_key, 0)
    for n in player_names:
        if (winner in grid_preds.get(n, {}).get(grid_key, set())
                and winner in actual):
            key = (n, winner, grid_key)
            if key not in awarded_grid:
                awarded_grid.add(key)
                earned[n] += pt_val
    return earned


def _build_daily_progression(matches, player_names, player_positions_pts=None,
                            all_groups_finished=False, grid_ctx=None):
    """Puntos acumulados tras cada partido jugado (orden cronológico), incluyendo fase final.

    Usa la misma lógica de puntuación que la tabla de clasificación:
    - Grupos: score del partido.
    - Eliminatoria: 1X2 + diferencia + exacto + bonus «clasificado correcto» (columna destino).
    - Cuadros de clasificados: q16 al acabar grupos; r8/r4/r2/final al clasificar cada equipo.
    """
    played_matches = [m for m in matches if m.get("played") and m.get("date")]
    if not played_matches:
        return {"labels": [], "flag_labels": [], "dates": [], "titles": [],
                "players": {n: [] for n in player_names},
                "day_points": {n: [] for n in player_names},
                "phases": []}

    grid_ctx = grid_ctx or {}
    player_q16_pts = grid_ctx.get("player_q16_pts") or {}
    grid_preds = grid_ctx.get("grid_preds") or {}
    pts_cfg = grid_ctx.get("pts") or {}
    actual_sets = grid_ctx.get("actual") or {}
    awarded_grid = set()

    played_matches.sort(key=lambda m: (m.get("date", ""), m.get("time_es", "")))

    cumulative  = {n: 0.0 for n in player_names}
    players_out = {n: [] for n in player_names}
    day_points  = {n: [] for n in player_names}
    labels      = []
    flag_labels = []
    dates  = []
    titles = []
    phases = []

    last_group_idx = -1
    if all_groups_finished and player_positions_pts:
        for i, m in enumerate(played_matches):
            if m["phase"] == "groups":
                last_group_idx = i

    def _match_title(m):
        gl, gv = m.get("goals_l"), m.get("goals_v")
        score = f" {gl}-{gv} " if gl is not None and gv is not None else " vs "
        title = f"{m.get('home','')}{score}{m.get('away','')}"
        if m.get("date"):
            try:
                dt = datetime.strptime(m["date"], "%Y-%m-%d")
                title += f" · {dt.day} {_MONTHS_ES[dt.month]}"
            except ValueError:
                pass
        return title

    for i, m in enumerate(played_matches):
        match_earned = {
            n: _prog_match_earned(m, m["predictions"].get(n, {}))
            for n in player_names
        }
        _append_prog_step(
            cumulative, players_out, day_points, labels, flag_labels,
            dates, titles, phases, player_names,
            f"{_abbr_team(m.get('home'))}-{_abbr_team(m.get('away'))}",
            f"{m.get('flag_home', '')}{m.get('flag_away', '')}" or f"{_abbr_team(m.get('home'))}-{_abbr_team(m.get('away'))}",
            m.get("date", ""),
            _match_title(m),
            m.get("phase", "groups"),
            match_earned,
        )

        if i == last_group_idx:
            _append_prog_step(
                cumulative, players_out, day_points, labels, flag_labels,
                dates, titles, phases, player_names,
                "Pos. Grupos", "🏆", dates[-1],
                "Posiciones de Fase de Grupos", "positions",
                {n: player_positions_pts.get(n, 0.0) for n in player_names},
            )
            if all_groups_finished and player_q16_pts:
                _append_prog_step(
                    cumulative, players_out, day_points, labels, flag_labels,
                    dates, titles, phases, player_names,
                    "Clasif. 16avos", "🏅", dates[-1],
                    "Clasificados a Dieciseisavos", "q16_team",
                    {n: player_q16_pts.get(n, 0.0) for n in player_names},
                )

        if m.get("phase") in ("r16", "r8", "r4", "r2") and grid_preds:
            grid_earned = _award_grid_on_ko_win(
                m, player_names, grid_preds, pts_cfg, awarded_grid, actual_sets,
            )
            if any(v > 0 for v in grid_earned.values()):
                w = str(m.get("actual_winner") or "").strip()
                abbr = _abbr_team(w) if w else "?"
                phase = m.get("phase")
                grid_phase = {"r16": "r8_team", "r8": "r4_team", "r4": "r2_team", "r2": "final_team"}.get(phase, "grid")
                _append_prog_step(
                    cumulative, players_out, day_points, labels, flag_labels,
                    dates, titles, phases, player_names,
                    f"Clasif. {abbr}", "✓", m.get("date", dates[-1] if dates else ""),
                    f"Clasificado a fase siguiente: {w}" if w else "Clasificado",
                    grid_phase,
                    grid_earned,
                )

    if all_groups_finished and player_positions_pts and last_group_idx == -1:
        last_date = dates[-1] if dates else datetime.now().strftime("%Y-%m-%d")
        _append_prog_step(
            cumulative, players_out, day_points, labels, flag_labels,
            dates, titles, phases, player_names,
            "Pos. Grupos", "🏆", last_date,
            "Posiciones de Fase de Grupos", "positions",
            {n: player_positions_pts.get(n, 0.0) for n in player_names},
        )
        if player_q16_pts:
            _append_prog_step(
                cumulative, players_out, day_points, labels, flag_labels,
                dates, titles, phases, player_names,
                "Clasif. 16avos", "🏅", last_date,
                "Clasificados a Dieciseisavos", "q16_team",
                {n: player_q16_pts.get(n, 0.0) for n in player_names},
            )

    progression = {
        "labels":      labels,
        "flag_labels": flag_labels,
        "dates":       dates,
        "titles":      titles,
        "players":     players_out,
        "day_points":  day_points,
        "phases":      phases,
    }
    return _sync_prog_players(progression, player_names)


def _load_file1():
    """Load 5 players from file [1]."""
    wb = openpyxl.load_workbook(FILE1, data_only=True)
    ws = wb["ADMIN"]
    players = []
    # col indices: (name_col, pred_col, score_col)
    defs = [
        ("S", 19, 19, 20),   # player letter, name_row5_col, pred_col, score_col
        ("V", 22, 22, 23),
        ("Y", 25, 25, 26),
        ("AB", 28, 28, 29),
        ("AE", 31, 31, 32),
    ]
    for letter, nc, pc, sc in defs:
        name = _val(ws, 5, nc)
        if not name or str(name).startswith("Pegar"):
            continue
        players.append({
            "name": str(name).strip(),
            "pred_col": pc,
            "score_col": sc,
        })

    clas_ws = wb["CLAS"]
    # read phase breakdowns from CLAS rows 5-9
    clas_data = {}
    for r in range(5, 10):
        player_name = _val(clas_ws, r, 3)
        if not player_name:
            continue
        clas_data[str(player_name).strip()] = {
            "total": _val(clas_ws, r, 4) or 0,
            "groups":    _val(clas_ws, r, 5)  or 0,
            "positions": _val(clas_ws, r, 6)  or 0,
            "q16":       _val(clas_ws, r, 7)  or 0,
            "r16":       _val(clas_ws, r, 8)  or 0,
            "r8":        _val(clas_ws, r, 9)  or 0,
            "r4":        _val(clas_ws, r, 10) or 0,
            "r2":        _val(clas_ws, r, 11) or 0,
            "r34_final": _val(clas_ws, r, 12) or 0,
            "honor":     _val(clas_ws, r, 13) or 0,
        }
    return ws, players, clas_data


def _load_file2():
    """Load Crespo from file [2]."""
    wb = openpyxl.load_workbook(FILE2, data_only=True)
    ws = wb["ADMIN"]
    name = _val(ws, 5, 19)
    if not name or str(name).startswith("Pegar"):
        name = "Crespo"
    player = {
        "name": str(name).strip(),
        "pred_col": 19,
        "score_col": 20,
    }
    clas_ws = wb["CLAS"]
    clas_data = {}
    for r in range(5, 10):
        player_name = _val(clas_ws, r, 3)
        if not player_name or str(player_name).startswith("Pegar"):
            continue
        clas_data[str(player_name).strip()] = {
            "total":     _val(clas_ws, r, 4)  or 0,
            "groups":    _val(clas_ws, r, 5)  or 0,
            "positions": _val(clas_ws, r, 6)  or 0,
            "q16":       _val(clas_ws, r, 7)  or 0,
            "r16":       _val(clas_ws, r, 8)  or 0,
            "r8":        _val(clas_ws, r, 9)  or 0,
            "r4":        _val(clas_ws, r, 10) or 0,
            "r2":        _val(clas_ws, r, 11) or 0,
            "r34_final": _val(clas_ws, r, 12) or 0,
            "honor":     _val(clas_ws, r, 13) or 0,
        }
    return ws, player, clas_data


def build_data():
    """Read both Excel files and return complete dashboard data."""
    global ADMIN, FILE1, FILE2
    ADMIN, FILE1, FILE2 = _excel_paths()

    wb1_raw = openpyxl.load_workbook(FILE1, data_only=True)
    ws1, players1, clas1 = _load_file1()
    ws2, player2, clas2  = _load_file2()

    spain_times = _build_spain_times(wb1_raw)
    wc_meta     = _build_wc_match_meta(wb1_raw)
    wc_scores   = _build_wc_scores(FILE1)   # reads from results.json (primary) or Excel fallback
    wc_scorers  = _load_scorers()
    wc_penalties = _load_penalties()
    wc_live     = _load_live()

    # ── results.json: authoritative match results + W-label resolution ──────────
    results_cache = _load_results_cache()
    ws_wc_for_j   = wb1_raw["WORLDCUP"]
    j_to_ah       = _build_j_to_ah(ws_wc_for_j)
    # wlabel_to_team: "W81" → "Estados Unidos", "L101" → "Marruecos", etc.
    wlabel_to_team = _build_wlabel_map(results_cache, j_to_ah)

    # Enrich wc_meta: add real-team-name keys for W-label entries now resolved
    for meta_key in list(wc_meta.keys()):
        mv = wc_meta[meta_key]
        h_raw = mv.get("home", "")
        a_raw = mv.get("away", "")
        if not (h_raw.startswith("W") or h_raw.startswith("L")):
            continue
        h_resolved = wlabel_to_team.get(h_raw)
        a_resolved = wlabel_to_team.get(a_raw)
        if h_resolved and a_resolved:
            real_key = f"{h_resolved}-{a_resolved}"
            if real_key not in wc_meta:
                new_mv = {**mv, "home": h_resolved, "away": a_resolved}
                wc_meta[real_key] = new_mv
                wc_meta[real_key.replace(" ", "")] = new_mv
            # Also add W-label scores to wc_scores so the lookup finds them
            ah_str = str(mv.get("match_num", ""))
            if ah_str in results_cache:
                entry = results_cache[ah_str]
                gl = entry.get("goals_h")
                gv = entry.get("goals_a")
                if gl is not None and gv is not None:
                    wc_scores[meta_key] = (int(gl), int(gv))
                    wc_scores[meta_key.replace(" ", "")] = (int(gl), int(gv))
    pts_sign  = float(_val(ws1, 8,  4) or 2)
    pts_diff  = float(_val(ws1, 9,  4) or 1)
    pts_exact = float(_val(ws1, 10, 4) or 3)
    diff_factor = float(_val(ws1, 50, 4) or 1)

    # ── Puntos de fases de eliminación (leídos del Excel, filas 16-38) ──
    KO_PHASE_PTS = {
        "r16":   {"sign": float(_val(ws1, 16, 4) or 3), "diff": float(_val(ws1, 17, 4) or 2), "exact": float(_val(ws1, 18, 4) or 4)},
        "r8":    {"sign": float(_val(ws1, 20, 4) or 4), "diff": float(_val(ws1, 21, 4) or 3), "exact": float(_val(ws1, 22, 4) or 5)},
        "r4":    {"sign": float(_val(ws1, 24, 4) or 5), "diff": float(_val(ws1, 25, 4) or 4), "exact": float(_val(ws1, 26, 4) or 6)},
        "r2":    {"sign": float(_val(ws1, 28, 4) or 6), "diff": float(_val(ws1, 29, 4) or 5), "exact": float(_val(ws1, 30, 4) or 8)},
        "r34":   {"sign": float(_val(ws1, 33, 4) or 6), "diff": float(_val(ws1, 34, 4) or 5), "exact": float(_val(ws1, 35, 4) or 8)},
        "final": {"sign": float(_val(ws1, 36, 4) or 8), "diff": float(_val(ws1, 37, 4) or 6), "exact": float(_val(ws1, 38, 4) or 12)},
    }
    # Mapeamos las fases del Excel (r34 y final) a las claves de standings (r34_final)
    _KO_STANDINGS_KEY = {"r16": "r16", "r8": "r8", "r4": "r4", "r2": "r2", "r34": "r34_final", "final": "r34_final"}

    _R8_BRACKET_MAP = {
        # match_num ADMIN → slot W del cuadro (home_placeholder-away_placeholder)
        89: {"name": "W73-W75", "home_num": 73, "away_num": 75},  # Canadá vs Marruecos
        90: {"name": "W74-W77", "home_num": 74, "away_num": 77},  # Paraguay vs Francia
        91: {"name": "W76-W78", "home_num": 76, "away_num": 78},
        92: {"name": "W79-W80", "home_num": 79, "away_num": 80},
        93: {"name": "W83-W84", "home_num": 83, "away_num": 84},
        94: {"name": "W81-W82", "home_num": 81, "away_num": 82},
        95: {"name": "W86-W88", "home_num": 86, "away_num": 88},
        96: {"name": "W85-W87", "home_num": 85, "away_num": 87},
    }

    # ── Cargar predicciones KO de los Excel individuales de cada jugador ──
    _ko_pred_path = os.path.join(BASE, "data", "ko_predictions.json")
    _ko_preds = {}
    if os.path.exists(_ko_pred_path):
        try:
            with open(_ko_pred_path, encoding="utf-8") as _f:
                _ko_preds = json.load(_f)
        except Exception:
            _ko_preds = {}

    all_players = players1 + [player2]
    all_ws      = [ws1] * len(players1) + [ws2]
    all_clas    = {**clas1, **clas2}

    player_names = [p["name"] for p in all_players]

    # ── Precalcular clasificados reales de cada fase para resolución de cruces ──
    def is_real_team(t):
        if not t: return False
        t_str = str(t).strip()
        if not t_str or t_str.startswith("1") or t_str.startswith("2") or t_str.startswith("3") or t_str.startswith("W") or t_str.startswith("L") or "-" in t_str:
            return False
        return True

    ws_wc = wb1_raw["WORLDCUP"]

    actual_q16_qualifiers = set()
    for r in range(101, 117):
        t1 = ws_wc.cell(r, 27).value
        t2 = ws_wc.cell(r, 32).value
        if is_real_team(t1): actual_q16_qualifiers.add(str(t1).strip())
        if is_real_team(t2): actual_q16_qualifiers.add(str(t2).strip())

    actual_r8_qualifiers = set()
    for r in range(120, 128):
        t1 = ws_wc.cell(r, 27).value
        t2 = ws_wc.cell(r, 32).value
        if is_real_team(t1): actual_r8_qualifiers.add(str(t1).strip())
        if is_real_team(t2): actual_r8_qualifiers.add(str(t2).strip())

    actual_r4_qualifiers = set()
    for r in range(131, 135):
        t1 = ws_wc.cell(r, 27).value
        t2 = ws_wc.cell(r, 32).value
        if is_real_team(t1): actual_r4_qualifiers.add(str(t1).strip())
        if is_real_team(t2): actual_r4_qualifiers.add(str(t2).strip())

    actual_r2_qualifiers = set()
    for r in range(138, 140):
        t1 = ws_wc.cell(r, 27).value
        t2 = ws_wc.cell(r, 32).value
        if is_real_team(t1): actual_r2_qualifiers.add(str(t1).strip())
        if is_real_team(t2): actual_r2_qualifiers.add(str(t2).strip())

    actual_r34_qualifiers = set()
    t1 = ws_wc.cell(143, 27).value
    t2 = ws_wc.cell(143, 32).value
    if is_real_team(t1): actual_r34_qualifiers.add(str(t1).strip())
    if is_real_team(t2): actual_r34_qualifiers.add(str(t2).strip())

    actual_final_qualifiers = set()
    t1 = ws_wc.cell(147, 27).value
    t2 = ws_wc.cell(147, 32).value
    if is_real_team(t1): actual_final_qualifiers.add(str(t1).strip())
    if is_real_team(t2): actual_final_qualifiers.add(str(t2).strip())

    # Reglas de puntos
    pts_q16_team = float(_val(ws1, 15, 4) or 2.0)
    pts_r8_team = float(_val(ws1, 19, 4) or 3.0)
    pts_r4_team = float(_val(ws1, 23, 4) or 5.0)
    pts_r2_team = float(_val(ws1, 27, 4) or 8.0)
    pts_r34_team = float(_val(ws1, 31, 4) or 8.0)
    pts_final_team = float(_val(ws1, 32, 4) or 12.0)
    actual_3rd_place = str(_val(ws1, 252, 13) or "").strip()
    actual_champion = str(_val(ws1, 250, 13) or "").strip()

    # Mapeo de selecciones a banderas detectadas
    team_to_flag = {}
    for k, v in wc_meta.items():
        h, a = v.get("home"), v.get("away")
        if h and v.get("flag_home") and not h.startswith("W") and not h.startswith("L") and not h.startswith("1") and not h.startswith("2"):
            team_to_flag[h] = v["flag_home"]
        if a and v.get("flag_away") and not a.startswith("W") and not a.startswith("L") and not a.startswith("1") and not a.startswith("2"):
            team_to_flag[a] = v["flag_away"]

    winner_by_num = {}
    match_by_num = {}

    # Pre-populate winner_by_num / match_by_num from results_cache so that
    # W-label resolve_team() works for all played matches even before the main
    # loop processes those rows.
    ah_to_j_map = {v: k for k, v in j_to_ah.items()}
    for _ah_str, _entry in results_cache.items():
        try:
            _ah = int(_ah_str)
        except ValueError:
            continue
        _h = _entry.get("home", "")
        _a = _entry.get("away", "")
        _w = _entry.get("winner", "")
        # winner_by_num uses AH match_num as key (consistent with wc_meta)
        if _h and _a:
            match_by_num[_ah] = {"home": _h, "away": _a}
        if _w:
            winner_by_num[_ah] = _w

    # ── collect all matches / prediction rows ────────────────────────────────
    matches = []
    played_count = {p["name"]: 0 for p in all_players}
    group_points = {p["name"]: 0.0 for p in all_players}
    ko_points    = {ph: {p["name"]: 0.0 for p in all_players} for ph in KO_PHASE_PTS}
    live_points  = {p["name"]: 0.0 for p in all_players}
    live_match_names = []
    spain_dates  = []

    for row in range(6, 268):
        match_name = _val(ws1, row, 11)  # K
        if not match_name or str(match_name).strip() in ("", "None"):
            continue

        phase = _phase_for_row(row)
        # Excluir filas que son etiquetas de sección o filas de clasificación del Excel
        name_str = str(match_name).strip()
        if phase in ("positions", "q16", "honor"):
            continue
        if "ENFRENTAMIENTOS" in name_str or "CLASIFICADOS" in name_str or "Equipos" in name_str or "Partidos" in name_str or "Partido Final" in name_str or "Partido 3-4" in name_str:
            continue
        # Evitar filas de marcadores/nombres provisionales genéricos que no son cruces reales
        if "Octavofinalista" in name_str or "Cuartofinalista" in name_str or "Semifinalista" in name_str or "Finalista" in name_str or "puesto-1" in name_str:
            continue
        match_id = _val(ws1, row, 10)    # J
        result_raw = _val(ws1, row, 13)  # M

        goals_l = _val(ws1, row, 15)   # O
        goals_v = _val(ws1, row, 16)   # P
        mkey = str(match_name).strip()

        # Resolve W/L-label keys (e.g. "W81-W82") to real team names if possible
        _wl_resolved = None
        if re.match(r'^[WL]\d+-[WL]\d+$', mkey):
            _parts = mkey.split("-", 1)
            _t0 = wlabel_to_team.get(_parts[0])
            _t1 = wlabel_to_team.get(_parts[1])
            if _t0 and _t1:
                _wl_resolved = f"{_t0}-{_t1}"
                # Update match_name / name_str for downstream lookups
                match_name = _wl_resolved
                name_str   = _wl_resolved

        # Score lookup: try original key, resolved name, and reversed resolved name
        _score_keys = [mkey]
        if _wl_resolved:
            _rev = "-".join(_wl_resolved.split("-", 1)[::-1])
            _score_keys += [_wl_resolved, _wl_resolved.replace(" ", ""), _rev, _rev.replace(" ", "")]
        _score_found = False
        for _sk in _score_keys:
            if _sk in wc_scores:
                goals_l, goals_v = wc_scores[_sk]
                result  = _result_from_goals(goals_l, goals_v)
                played  = result is not None
                _score_found = True
                break
        if not _score_found:
            result = _parse_result(result_raw)
            played = result is not None
        mult_raw = _val(ws1, row, 9)
        try:
            multiplier = float(mult_raw) if mult_raw is not None else 1.0
        except (ValueError, TypeError):
            multiplier = 1.0

        # Spain datetime
        spain_dt = _lookup_spain_time(spain_times, match_name)
        if spain_dt:
            spain_dates.append(spain_dt)
            date_es  = spain_dt.strftime("%Y-%m-%d")
            time_es  = spain_dt.strftime("%H:%M")
            day_label = spain_dt.strftime("%A %d %B").replace(
                "Monday", "Lunes").replace("Tuesday", "Martes").replace(
                "Wednesday", "Miércoles").replace("Thursday", "Jueves").replace(
                "Friday", "Viernes").replace("Saturday", "Sábado").replace(
                "Sunday", "Domingo")
            # Spanish month names
            for en, es in [("January","enero"),("February","febrero"),("March","marzo"),
                           ("April","abril"),("May","mayo"),("June","junio"),
                           ("July","julio"),("August","agosto"),("September","septiembre"),
                           ("October","octubre"),("November","noviembre"),("December","diciembre")]:
                day_label = day_label.replace(en, es)
        else:
            h_val = _val(ws1, row, 8)
            date_es = str(h_val)[:10] if h_val else ""
            time_es = ""
            day_label = date_es
            # Fallback: horario en fixture_data cuando Excel no tiene fecha válida
            if not date_es.startswith("2026-"):
                fix_sched = lookup_fixture(row)
                if fix_sched.get("date", "").startswith("2026-"):
                    date_es   = fix_sched["date"]
                    time_es   = fix_sched.get("time_es", "")
                    day_label = fix_sched.get("day_label", date_es)
                    try:
                        spain_dt = datetime.strptime(
                            f"{date_es} {time_es or '00:00'}", "%Y-%m-%d %H:%M"
                        )
                    except ValueError:
                        spain_dt = None

        predictions = {}
        # ── overlay EN CURSO: solo si el partido no está finalizado en Excel ──
        live_info = None
        if not played:
            li = _lookup_live(wc_live, match_name)
            if li is not None:
                try:
                    lgl = int(li.get("home"))
                    lgv = int(li.get("away"))
                except (TypeError, ValueError):
                    lgl = lgv = None
                if lgl is not None and lgv is not None:
                    live_info = {
                        "goals_l": lgl,
                        "goals_v": lgv,
                        "minute":  str(li.get("minute", "")).strip(),
                        "result":  _result_from_goals(lgl, lgv),
                        "scorers": li.get("scorers", []),
                    }
                    live_match_names.append(str(match_name).strip())

        wc = _lookup_wc_meta(wc_meta, match_name) or {}  # needed before player loop for team matching
        
        home_raw = wc.get("home", "")
        away_raw = wc.get("away", "")
        
        def resolve_team(ph):
            if not ph: return ph, ""
            ph_str = str(ph).strip()
            if ph_str.startswith("W") and ph_str[1:].isdigit():
                num = int(ph_str[1:])
                w = winner_by_num.get(num)
                if w: return w, team_to_flag.get(w, "")
            if ph_str.startswith("L") and ph_str[1:].isdigit():
                num = int(ph_str[1:])
                w = winner_by_num.get(num)
                m_orig = match_by_num.get(num)
                if w and m_orig:
                    h_orig = m_orig.get("home")
                    a_orig = m_orig.get("away")
                    if h_orig and a_orig and not h_orig.startswith("W") and not h_orig.startswith("L") and not a_orig.startswith("W") and not a_orig.startswith("L"):
                        loser = a_orig if h_orig == w else h_orig
                        return loser, team_to_flag.get(loser, "")
            return ph_str, ""

        home_resolved, flag_h = resolve_team(home_raw)
        if not flag_h: flag_h = wc.get("flag_home", "")
        away_resolved, flag_a = resolve_team(away_raw)
        if not flag_a: flag_a = wc.get("flag_away", "")

        wc["home"] = home_resolved
        wc["flag_home"] = flag_h
        wc["away"] = away_resolved
        wc["flag_away"] = flag_a

        # Reintentar marcador con nombres ya resueltos (p. ej. ADMIN «Paraguay-Francia»)
        if not played and home_resolved and away_resolved:
            for sk in (f"{home_resolved}-{away_resolved}",
                       f"{away_resolved}-{home_resolved}"):
                if sk in wc_scores:
                    goals_l, goals_v = wc_scores[sk]
                    result = _result_from_goals(goals_l, goals_v)
                    played = result is not None
                    break

        actual_home_set = bool(home_resolved and not home_resolved.startswith("1") and not home_resolved.startswith("2") and not home_resolved.startswith("W") and not home_resolved.startswith("L"))
        actual_away_set = bool(away_resolved and not away_resolved.startswith("1") and not away_resolved.startswith("2") and not away_resolved.startswith("W") and not away_resolved.startswith("L"))
        
        for p, ws in zip(all_players, all_ws):
            if phase in KO_PHASE_PTS:
                # Lookup order differs by phase:
                #   r8: bracket slot (W-label from placeholders) first — this is how
                #       players store predictions in their Excel (e.g. "W74-W77").
                #       Row/match_num differ between player Excels and ADMIN, so they
                #       must NOT be used as primary keys.
                #   r16 and others: name first, because player Excels number rows
                #       differently from ADMIN but team-name keys are correct.
                m_num = wc.get("match_num")
                name_str = str(match_name).strip()
                ko_player = _ko_preds.get(p["name"], {})
                pred_raw = None
                h_ph = str(wc.get("home_placeholder") or "").strip()
                a_ph = str(wc.get("away_placeholder") or "").strip()
                bracket_slot = f"{h_ph}-{a_ph}" if h_ph.startswith("W") and a_ph.startswith("W") else None
                if phase == "r8":
                    key_order = [
                        bracket_slot,
                        _R8_BRACKET_MAP[m_num]["name"] if m_num in _R8_BRACKET_MAP else None,
                        name_str,
                        f"{home_resolved}-{away_resolved}" if actual_home_set and actual_away_set else None,
                        f"{away_resolved}-{home_resolved}" if actual_home_set and actual_away_set else None,
                        str(m_num) if m_num is not None else None,
                        str(row),
                        mkey,
                    ]
                else:
                    key_order = [
                        name_str,
                        str(row),
                        str(m_num) if m_num is not None else None,
                        f"{home_resolved}-{away_resolved}" if actual_home_set and actual_away_set else None,
                        f"{away_resolved}-{home_resolved}" if actual_home_set and actual_away_set else None,
                        _R8_BRACKET_MAP[m_num]["name"] if phase == "r8" and m_num in _R8_BRACKET_MAP else None,
                        mkey,
                    ]
                for key in filter(None, key_order):
                    pred_raw = ko_player.get(key)
                    if pred_raw:
                        break
                score_raw = 0 # Score will be calculated below
            else:
                pred_raw  = _val(ws, row, p["pred_col"])
                score_raw = _val(ws, row, p["score_col"])
            
            pred  = _parse_pred(pred_raw)
            score = float(score_raw) if score_raw is not None else 0

            breakdown = None
            live_breakdown = None
            if played and phase == "groups" and pred and "|" in str(pred_raw or ""):
                gl = int(goals_l) if goals_l is not None else None
                gv = int(goals_v) if goals_v is not None else None
                breakdown = _score_breakdown(
                    pred, result, gl, gv,
                    pts_sign, pts_diff, pts_exact,
                    diff_factor, multiplier,
                )
                # El punto calculado en Python manda sobre la fórmula del Excel
                # (que puede estar desactualizada tras una actualización automática).
                score = breakdown["total"]
                group_points[p["name"]] += breakdown["total"]
            elif played and phase in KO_PHASE_PTS and pred and "|" in str(pred_raw or ""):
                # Fases de eliminación: puntos calculados en Python igual que grupos.
                # Primero verificamos si el jugador prediøjo los equipos correctos.
                gl = int(goals_l) if goals_l is not None else None
                gv = int(goals_v) if goals_v is not None else None
                actual_home = wc.get("home", "")
                actual_away = wc.get("away", "")
                ph = pred.get("pred_home", "")
                pa = pred.get("pred_away", "")
                home_ok      = bool(ph and actual_home and ph.strip() == actual_home.strip())
                away_ok      = bool(pa and actual_away and pa.strip() == actual_away.strip())
                # Detectar equipos puestos en orden invertido (ej. predijo Francia-Marruecos
                # pero el cruce real es Marruecos-Francia). La regla habla de acertar
                # los DOS equipos, sin importar si los puso como local o visitante.
                home_as_away = bool(ph and actual_away and ph.strip() == actual_away.strip())
                away_as_home = bool(pa and actual_home and pa.strip() == actual_home.strip())
                teams_reversed = home_as_away and away_as_home
                # team_match: 'both', 'home', 'away', 'none', or None if no team pred
                if ph and pa:
                    if (home_ok and away_ok) or teams_reversed:
                        team_match = "both"
                    elif home_ok or home_as_away:
                        team_match = "home" if home_ok else "away"
                    elif away_ok or away_as_home:
                        team_match = "away" if away_ok else "home"
                    else:
                        team_match = "none"
                else:
                    team_match = None
                if phase == "r16":
                    # En 16avos no hay penalización por equipos incorrectos
                    team_match = None
                ko_cfg = KO_PHASE_PTS[phase]
                # Si los equipos están invertidos, flipear goles y sign para la
                # comparación (el jugador apostó «local»/«visitante» al revés)
                _cmp_gl, _cmp_gv, _cmp_result = gl, gv, result
                if teams_reversed and not (home_ok and away_ok) and gl is not None and gv is not None:
                    _cmp_gl, _cmp_gv = gv, gl
                    if _cmp_result:
                        _flipped_sign = ("2" if _cmp_result.get("sign") == "1"
                                         else ("1" if _cmp_result.get("sign") == "2"
                                               else _cmp_result.get("sign", "")))
                        _cmp_result = {**_cmp_result, "sign": _flipped_sign,
                                       "score": f"{_cmp_gl}-{_cmp_gv}"}
                breakdown = _score_breakdown(
                    pred, _cmp_result, _cmp_gl, _cmp_gv,
                    ko_cfg["sign"], ko_cfg["diff"], ko_cfg["exact"],
                    diff_factor, multiplier,
                )
                # REGLA DE ORO: para puntuar por resultado (1X2/diff/exacto) hay
                # que acertar LOS DOS equipos. Con solo uno correcto (o ninguno)
                # los puntos de resultado son 0; la excepción del "Clasificado"
                # se aplica más abajo si ese único equipo acertado avanza.
                if phase != "r16" and team_match != "both" and team_match is not None:
                    if team_match == "none":
                        _reason = "Equipos incorrectos (0 pts)"
                    else:
                        _reason = "Solo un equipo correcto — sin puntos de resultado"
                    breakdown = {**breakdown, "total": 0.0, "sign": 0.0, "diff": 0.0, "exact": 0.0,
                                 "reasons": [_reason]}
                breakdown["team_match"] = team_match
                breakdown["pred_home"] = ph
                breakdown["pred_away"] = pa
                score = breakdown["total"]
                ko_points[phase][p["name"]] += breakdown["total"]

            elif live_info and phase == "groups" and pred and "|" in str(pred_raw or ""):
                # Puntos provisionales del partido en curso (no oficiales).
                live_breakdown = _score_breakdown(
                    pred, live_info["result"],
                    live_info["goals_l"], live_info["goals_v"],
                    pts_sign, pts_diff, pts_exact,
                    diff_factor, multiplier,
                )
                live_points[p["name"]] += live_breakdown["total"]
            elif live_info and phase in KO_PHASE_PTS and pred and "|" in str(pred_raw or ""):
                # Puntos provisionales EN CURSO para fases KO.
                actual_home = wc.get("home", "")
                actual_away = wc.get("away", "")
                ph = pred.get("pred_home", "")
                pa = pred.get("pred_away", "")
                home_ok      = bool(ph and actual_home and ph.strip() == actual_home.strip())
                away_ok      = bool(pa and actual_away and pa.strip() == actual_away.strip())
                home_as_away = bool(ph and actual_away and ph.strip() == actual_away.strip())
                away_as_home = bool(pa and actual_home and pa.strip() == actual_home.strip())
                teams_reversed = home_as_away and away_as_home
                if ph and pa:
                    if (home_ok and away_ok) or teams_reversed:
                        team_match = "both"
                    elif home_ok or home_as_away:
                        team_match = "home" if home_ok else "away"
                    elif away_ok or away_as_home:
                        team_match = "away" if away_ok else "home"
                    else:
                        team_match = "none"
                else:
                    team_match = None
                if phase == "r16":
                    team_match = None
                ko_cfg = KO_PHASE_PTS[phase]
                _lgl = live_info["goals_l"]; _lgv = live_info["goals_v"]
                _lresult = live_info["result"]
                if teams_reversed and not (home_ok and away_ok) and _lgl is not None and _lgv is not None:
                    _lgl, _lgv = _lgv, _lgl
                    if _lresult:
                        _flipped_sign = ("2" if _lresult.get("sign") == "1"
                                         else ("1" if _lresult.get("sign") == "2"
                                               else _lresult.get("sign", "")))
                        _lresult = {**_lresult, "sign": _flipped_sign, "score": f"{_lgl}-{_lgv}"}
                live_breakdown = _score_breakdown(
                    pred, _lresult, _lgl, _lgv,
                    ko_cfg["sign"], ko_cfg["diff"], ko_cfg["exact"],
                    diff_factor, multiplier,
                )
                if phase != "r16" and team_match != "both" and team_match is not None:
                    if team_match == "none":
                        _reason = "Equipos incorrectos (0 pts)"
                    else:
                        _reason = "Solo un equipo correcto — sin puntos de resultado"
                    live_breakdown = {**live_breakdown, "total": 0.0, "sign": 0.0, "diff": 0.0, "exact": 0.0,
                                      "reasons": [_reason]}
                live_breakdown["team_match"] = team_match
                live_breakdown["pred_home"] = ph
                live_breakdown["pred_away"] = pa
                live_points[p["name"]] += live_breakdown["total"]


            predictions[p["name"]] = {
                "pred":      pred,
                "score":     score,
                "breakdown": breakdown,
                "live_breakdown": live_breakdown,
                "live_score":     live_breakdown["total"] if live_breakdown else None,
            }

        if phase == "groups" and played:
            for name in player_names:
                played_count[name] += 1

        fix = lookup_fixture(row)
        r16_order = [73, 75, 74, 77, 76, 78, 79, 80, 83, 84, 81, 82, 86, 88, 85, 87]
        r8_order = [200, 201, 204, 205, 202, 203, 206, 207]
        r4_order = [220, 221, 222, 223]
        r2_order = [232, 233]
        b_order = 0
        if phase == "r16" and wc.get("match_num") in r16_order: b_order = r16_order.index(wc.get("match_num"))
        elif phase == "r8" and row in r8_order: b_order = r8_order.index(row)
        elif phase == "r4" and row in r4_order: b_order = r4_order.index(row)
        elif phase == "r2" and row in r2_order: b_order = r2_order.index(row)

        matches.append({
            "row":       row,
            "id":        str(match_id) if match_id else "",
            "name":      str(match_name).strip(),
            "home":      wc.get("home", ""),
            "away":      wc.get("away", ""),
            "home_placeholder": wc.get("home_placeholder", ""),
            "away_placeholder": wc.get("away_placeholder", ""),
            "flag_home": wc.get("flag_home", ""),
            "flag_away": wc.get("flag_away", ""),
            "match_num": wc.get("match_num"),
            "bracket_order": b_order,
            "city":       fix.get("city", ""),
            "country":    fix.get("country", ""),
            "tv":         "both" if (wc.get("home", "") == "España" or wc.get("away", "") == "España") else fix.get("tv", ""),
            "tv_label":   TV_LABELS.get("both" if (wc.get("home", "") == "España" or wc.get("away", "") == "España") else fix.get("tv", ""), ""),
            "stadium":    fix.get("stadium", ""),
            "lat":        fix.get("lat"),
            "lon":        fix.get("lon"),
            "capacity":   fix.get("capacity"),
            "city_pop":   fix.get("city_pop", ""),
            "venue_fact": fix.get("fact", ""),
            "wiki":       fix.get("wiki", ""),
            "date":      date_es,
            "time_es":   time_es,
            "day_label": day_label,
            "datetime_es": spain_dt.isoformat() if spain_dt else date_es,
            "phase":     phase,
            "result":    result,
            "played":    played,
            "live":      bool(live_info),
            "live_minute":  live_info["minute"] if live_info else "",
            "live_goals_l": live_info["goals_l"] if live_info else None,
            "live_goals_v": live_info["goals_v"] if live_info else None,
            "live_scorers": live_info["scorers"] if live_info else [],
            "goals_l":   int(goals_l) if (played and goals_l is not None) else None,
            "goals_v":   int(goals_v) if (played and goals_v is not None) else None,
            "scorers":   (_lookup_scorers(wc_scorers, match_name)
                          or _lookup_scorers(wc_scorers, f"{home_resolved}-{away_resolved}")
                          or _lookup_scorers(wc_scorers, f"{away_resolved}-{home_resolved}"))
                         if played else [],
            "phase_pts":  KO_PHASE_PTS.get(phase),  # puntos posibles por criterio para este partido
            "predictions": predictions,
        })

        # Determine actual winner for this match on the fly
        m_obj = matches[-1]
        m_obj["penalties"] = wc_penalties.get(match_name)
        m_obj["actual_winner"] = None
        if played and home_resolved and away_resolved:
            gl = m_obj["goals_l"]
            gv = m_obj["goals_v"]
            if gl is not None and gv is not None:
                if gl > gv: m_obj["actual_winner"] = home_resolved
                elif gv > gl: m_obj["actual_winner"] = away_resolved
                else:
                    p_info = wc_penalties.get(match_name)
                    if p_info:
                        if p_info["home"] > p_info["away"]: m_obj["actual_winner"] = home_resolved
                        elif p_info["away"] > p_info["home"]: m_obj["actual_winner"] = away_resolved

                    if not m_obj["actual_winner"]:
                        if phase == "r16" and home_resolved in actual_r8_qualifiers: m_obj["actual_winner"] = home_resolved
                        elif phase == "r16" and away_resolved in actual_r8_qualifiers: m_obj["actual_winner"] = away_resolved
                        elif phase == "r8" and home_resolved in actual_r4_qualifiers: m_obj["actual_winner"] = home_resolved
                        elif phase == "r8" and away_resolved in actual_r4_qualifiers: m_obj["actual_winner"] = away_resolved
                        elif phase == "r4" and home_resolved in actual_r2_qualifiers: m_obj["actual_winner"] = home_resolved
                        elif phase == "r4" and away_resolved in actual_r2_qualifiers: m_obj["actual_winner"] = away_resolved
                        elif phase == "r2" and home_resolved in actual_final_qualifiers: m_obj["actual_winner"] = home_resolved
                        elif phase == "r2" and away_resolved in actual_final_qualifiers: m_obj["actual_winner"] = away_resolved
                        elif phase == "r34" and actual_3rd_place in (home_resolved, away_resolved): m_obj["actual_winner"] = actual_3rd_place
                        elif phase == "final" and actual_champion in (home_resolved, away_resolved): m_obj["actual_winner"] = actual_champion
        
        # Save to maps for Wxx / Lxx resolution in future rows
        match_num = wc.get("match_num")
        if match_num:
            match_by_num[match_num] = {"home": home_resolved, "away": away_resolved}
            if m_obj["actual_winner"]:
                winner_by_num[match_num] = m_obj["actual_winner"]

    _backfill_match_flags(matches)
    _propagate_bracket_winners(matches)

    # ── Corregir bracket_order visual de los octavos ──────────────────────────
    # En el Admin Excel los octavos van en pares: (89,90), (91,92), (93,94), (95,96).
    # El primer partido de cada par (ej. 89) alimenta visualmente desde los 16avos
    # de abajo (Alemania/Paraguay + Francia/Suecia), pero bracket_order=0 lo pone
    # arriba. Intercambiamos los bracket_order dentro de cada par para que el
    # partido cuyas fuentes son los 16avos superiores quede arriba visualmente.
    _R8_ORDER_SWAP_PAIRS = [(89, 90), (91, 92), (93, 94), (95, 96)]
    for _mn_a, _mn_b in _R8_ORDER_SWAP_PAIRS:
        _ma = next((x for x in matches if x.get("match_num") == _mn_a), None)
        _mb = next((x for x in matches if x.get("match_num") == _mn_b), None)
        if _ma and _mb:
            _ma["bracket_order"], _mb["bracket_order"] = _mb["bracket_order"], _ma["bracket_order"]

    _augment_actual_qualifiers_from_matches(
        matches, actual_r8_qualifiers, actual_r4_qualifiers, actual_r2_qualifiers,
        actual_r34_qualifiers, actual_final_qualifiers,
    )

    # Para evitar depender de la caché de fórmulas de Excel, que no se actualiza
    # al escribir datos programáticamente.
    actual_standings = {}
    group_matches_count = {}
    for m in matches:
        if m["phase"] == "groups" and m["id"]:
            grp = m["id"][0]
            group_matches_count[grp] = group_matches_count.get(grp, 0) + (1 if m["played"] else 0)

    # Calcular clasificaciones de grupos en base a partidos
    group_teams = {}
    for m in matches:
        if m["phase"] == "groups" and m["id"]:
            grp = m["id"][0]
            if grp not in group_teams:
                group_teams[grp] = {}
            for t in (m["home"], m["away"]):
                if t and t not in group_teams[grp]:
                    group_teams[grp][t] = {"pts": 0, "gd": 0, "gf": 0, "gc": 0, "name": t}
            if m["played"] and m["goals_l"] is not None and m["goals_v"] is not None:
                h, a = m["home"], m["away"]
                gl, gv = m["goals_l"], m["goals_v"]
                group_teams[grp][h]["gf"] += gl
                group_teams[grp][h]["gc"] = group_teams[grp][h].get("gc", 0) + gv
                group_teams[grp][a]["gf"] += gv
                group_teams[grp][a]["gc"] = group_teams[grp][a].get("gc", 0) + gl
                if gl > gv:
                    group_teams[grp][h]["pts"] += 3
                elif gl < gv:
                    group_teams[grp][a]["pts"] += 3
                else:
                    group_teams[grp][h]["pts"] += 1
                    group_teams[grp][a]["pts"] += 1

    for grp, teams in group_teams.items():
        for t in teams.values():
            t["gd"] = t["gf"] - t["gc"]
        sorted_teams = sorted(teams.values(), key=lambda x: (x["pts"], x["gd"], x["gf"]), reverse=True)
        actual_standings[grp] = [t["name"] for t in sorted_teams]

    actual_positions_map = {}
    pts_pos_rules = {
        0: float(_val(ws1, 11, 4) or 4.0), # 1º
        1: float(_val(ws1, 12, 4) or 3.0), # 2º
        2: float(_val(ws1, 13, 4) or 2.0), # 3º
        3: float(_val(ws1, 14, 4) or 1.0)  # 4º
    }

    total_group_matches_played = sum(1 for m in matches if m["phase"] == "groups" and m["played"])
    all_groups_finished = (total_group_matches_played == 72)

    for r in range(80, 128):
        grp = _val(ws1, r, 10)
        if not grp:
            k_val = str(_val(ws1, r, 11))
            m_g = re.search(r'GRUPO\s+([A-L])', k_val)
            grp = m_g.group(1) if m_g else None
        pos_idx = (r - 80) % 4
        if grp and all_groups_finished:
            actual_positions_map[r] = actual_standings[grp][pos_idx]
        else:
            actual_positions_map[r] = f"{pos_idx+1}{grp}"

    player_positions_pts = {}
    player_q16_pts = {}
    player_r8_team_pts = {}
    player_r4_team_pts = {}
    player_r2_team_pts = {}
    player_r34_team_pts = {}
    player_final_team_pts = {}

    for p, ws in zip(all_players, all_ws):
        name = p["name"]
        pos_pts = 0.0
        for r in range(80, 128):
            actual = actual_positions_map.get(r)
            pred = _val(ws, r, p["pred_col"])
            if actual and pred and str(actual).strip() == str(pred).strip():
                pos_pts += pts_pos_rules[(r - 80) % 4]
        player_positions_pts[name] = pos_pts

        # q16_team (Clasificados 16avos)
        q16_pts = 0.0
        for r in range(130, 162):
            pred = _val(ws, r, p["pred_col"])
            if pred and str(pred).strip() in actual_q16_qualifiers:
                q16_pts += pts_q16_team
        player_q16_pts[name] = q16_pts

        # r8_team (Clasificados Octavos)
        r8_team_pts = 0.0
        for r in range(182, 198):
            pred = _val(ws, r, p["pred_col"])
            if pred and str(pred).strip() in actual_r8_qualifiers:
                r8_team_pts += pts_r8_team
        player_r8_team_pts[name] = r8_team_pts

        # r4_team (Clasificados Cuartos)
        r4_team_pts = 0.0
        for r in range(210, 218):
            pred = _val(ws, r, p["pred_col"])
            if pred and str(pred).strip() in actual_r4_qualifiers:
                r4_team_pts += pts_r4_team
        player_r4_team_pts[name] = r4_team_pts

        # r2_team (Clasificados Semifinales)
        r2_team_pts = 0.0
        for r in range(226, 230):
            pred = _val(ws, r, p["pred_col"])
            if pred and str(pred).strip() in actual_r2_qualifiers:
                r2_team_pts += pts_r2_team
        player_r2_team_pts[name] = r2_team_pts

        # r34_team (Clasificados 3º y 4º Puesto)
        r34_team_pts = 0.0
        for r in range(236, 238):
            pred = _val(ws, r, p["pred_col"])
            if pred and str(pred).strip() in actual_r34_qualifiers:
                r34_team_pts += pts_r34_team
        player_r34_team_pts[name] = r34_team_pts

        # final_team (Clasificados Final)
        final_team_pts = 0.0
        for r in range(240, 242):
            pred = _val(ws, r, p["pred_col"])
            if pred and str(pred).strip() in actual_final_qualifiers:
                final_team_pts += pts_final_team
        player_final_team_pts[name] = final_team_pts

    player_qual_match_pts = {
        "r8": {n: 0.0 for n in player_names},
        "r4": {n: 0.0 for n in player_names},
        "r2": {n: 0.0 for n in player_names},
        "r34_final": {n: 0.0 for n in player_names},
    }

    # ── Sumar puntos de clasificados a las predicciones individuales de cada partido ──
    for m in matches:
        phase = m["phase"]
        if phase in ("r16", "r8", "r4", "r2", "r34", "final") and m["played"]:
            actual_w = m["actual_winner"]
            if not actual_w:
                continue
            qual_pts = 0.0
            if phase == "r16": qual_pts = pts_r8_team
            elif phase == "r8": qual_pts = pts_r4_team
            elif phase == "r4": qual_pts = pts_r2_team
            
            for p in all_players:
                name = p["name"]
                pred_obj = m["predictions"].get(name)
                if pred_obj and pred_obj.get("pred"):
                    # El bonus de clasificado solo aplica si se acertó al menos un equipo
                    # (team_match 'both', 'home' o 'away'); con team_match 'none' no hay excepción
                    _tm = (pred_obj.get("breakdown") or {}).get("team_match")
                    if _tm == "none":
                        continue
                    pred_w = pred_obj["pred"].get("winner")
                    if pred_w and str(pred_w).strip().lower() == str(actual_w).strip().lower():
                        this_pts = qual_pts
                        if phase == "r2":
                            player_ws = all_ws[player_names.index(name)]
                            player_idx = player_names.index(name)
                            pred_col = all_players[player_idx]["pred_col"]
                            predicted_finalists = [str(_val(player_ws, r, pred_col) or "").strip().lower() for r in (240, 241)]
                            if str(actual_w).strip().lower() in predicted_finalists:
                                this_pts = pts_final_team
                            else:
                                this_pts = pts_r34_team
                        elif phase == "r34":
                            this_pts = pts_r34_team
                        elif phase == "final":
                            this_pts = pts_final_team
                            
                        pred_obj["score"] = (pred_obj.get("score") or 0.0) + this_pts
                        pred_obj["qual_pts"] = (pred_obj.get("qual_pts") or 0.0) + this_pts
                        if pred_obj.get("breakdown"):
                            pred_obj["breakdown"]["total"] += this_pts
                            pred_obj["breakdown"]["reasons"].append(f"Clasificado correcto (+{this_pts} pts)")
                        target = _ko_match_qual_target_phase(m)
                        if target and target in player_qual_match_pts:
                            player_qual_match_pts[target][name] += this_pts

    # ── standings ────────────────────────────────────────────────────────────
    standings_raw = []
    for i, p in enumerate(all_players):
        name = p["name"]
        clas = all_clas.get(name, {})

        def fv(key):
            v = clas.get(key, 0)
            try: return float(v) if v else 0.0
            except: return 0.0

        groups_calc = round(group_points.get(name, 0.0), 2)
        groups_excel = fv("groups")
        
        positions_calc = player_positions_pts.get(name, 0.0)
        positions_excel = fv("positions")
        
        q16_calc      = player_q16_pts.get(name, 0.0)
        r16_calc      = round(ko_points["r16"].get(name, 0.0) + q16_calc, 2)
        r8_calc       = round(ko_points["r8"].get(name, 0.0) + player_r8_team_pts.get(name, 0.0)
                          + player_qual_match_pts["r8"].get(name, 0.0), 2)
        r4_calc       = round(ko_points["r4"].get(name, 0.0) + player_r4_team_pts.get(name, 0.0)
                          + player_qual_match_pts["r4"].get(name, 0.0), 2)
        r2_calc       = round(ko_points["r2"].get(name, 0.0) + player_r2_team_pts.get(name, 0.0)
                          + player_qual_match_pts["r2"].get(name, 0.0), 2)
        r34_final_calc = round(
            ko_points["r34"].get(name, 0.0) + ko_points["final"].get(name, 0.0)
            + player_r34_team_pts.get(name, 0.0) + player_final_team_pts.get(name, 0.0)
            + player_qual_match_pts["r34_final"].get(name, 0.0),
            2
        )

        total = round(groups_calc + positions_calc + r16_calc + r8_calc + r4_calc + r2_calc + r34_final_calc, 2)
        lp = round(live_points.get(name, 0.0), 2)
        total_live = round(total + lp, 2)
        phase_detail = []
        for key, label, desc in STANDINGS_PHASES:
            if key == "groups":
                pts = groups_calc
            elif key == "positions":
                pts = positions_calc
            elif key == "r16":
                pts = r16_calc
            elif key == "r8":
                pts = r8_calc
            elif key == "r4":
                pts = r4_calc
            elif key == "r2":
                pts = r2_calc
            elif key == "r34_final":
                pts = r34_final_calc
            else:
                pts = 0.0
            if pts > 0:
                phase_detail.append({"key": key, "label": label, "desc": desc, "pts": pts})

        standings_raw.append({
            "name":     name,
            "total":    total,
            "live_points": lp,
            "total_live":  total_live,
            "groups":   groups_calc,
            "positions": positions_calc,
            "q16":      q16_calc,
            "r16":      r16_calc,
            "r8":       r8_calc,
            "r4":       r4_calc,
            "r2":       r2_calc,
            "r34_final":r34_final_calc,
            "honor":    0.0,
            "phase_detail": phase_detail,
            "color":    PLAYER_COLORS[i % len(PLAYER_COLORS)],
            "played":   played_count.get(name, 0),
            "_excel_total": fv("total"),  # total del Excel (jornada anterior)
            "_orig_idx": i,               # orden original para desempate
        })

    # ── Posiciones previas para desempate y flechas de cambio ────────────────
    # prev_standings.json almacena 3 campos:
    #   display_positions — posiciones de ANTES del último cambio (para flechas ▲▼=)
    #   current_positions — posiciones tras el último cambio (tiebreak + futuro display)
    #   totals            — totales tras el último cambio (detectar nuevo cambio)
    prev_standings_path = os.path.join(BASE, "data", "prev_standings.json")
    display_pos_map = {}   # para mostrar flechas
    tiebreak_pos_map = {}  # para desempate en sort
    prev_totals = {}
    try:
        if os.path.isfile(prev_standings_path):
            with open(prev_standings_path, encoding="utf-8") as f:
                prev_data = json.load(f)
            display_pos_map = prev_data.get("display_positions",
                                            prev_data.get("positions", {}))
            tiebreak_pos_map = prev_data.get("current_positions",
                                             prev_data.get("positions", {}))
            prev_totals = prev_data.get("totals", {})
    except (OSError, ValueError):
        pass

    # Desempate: si dos jugadores tienen los mismos puntos, el que tenía
    # mejor posición previa (pos más baja) se mantiene por encima.
    # Usa current_positions (las más recientes confirmadas) para tiebreak.
    standings_raw.sort(
        key=lambda x: (
            x["total"],
            -tiebreak_pos_map.get(x["name"], 99),   # menor pos previa = mejor
            x.get("_excel_total", 0),
            -x.get("_orig_idx", 0),
        ),
        reverse=True,
    )

    # Determinar posiciones actuales y ver si hubo cambio de puntos
    current_positions = {}
    current_totals = {}
    for i, s in enumerate(standings_raw):
        current_positions[s["name"]] = i + 1
        current_totals[s["name"]] = s["total"]

    # Si los puntos cambiaron respecto al prev guardado, actualizamos prev
    if current_totals != prev_totals and prev_totals:
        # Los puntos han cambiado:
        #   display_positions ← las posiciones que había JUSTO ANTES (current del fichero)
        #   current_positions ← las posiciones NUEVAS tras el cambio
        #   totals ← los totales nuevos
        try:
            with open(prev_standings_path, "w", encoding="utf-8") as f:
                json.dump({
                    "display_positions": tiebreak_pos_map,
                    "current_positions": current_positions,
                    "totals": current_totals,
                }, f, ensure_ascii=False, indent=2)
        except OSError:
            pass
        display_pos_map = tiebreak_pos_map  # flechas comparan contra el estado anterior
    elif not prev_totals:
        # Primera vez (no hay prev) → guardar estado actual, sin flechas
        display_pos_map = current_positions
        try:
            with open(prev_standings_path, "w", encoding="utf-8") as f:
                json.dump({
                    "display_positions": current_positions,
                    "current_positions": current_positions,
                    "totals": current_totals,
                }, f, ensure_ascii=False, indent=2)
        except OSError:
            pass

    standings = []
    for i, s in enumerate(standings_raw):
        s["pos"] = i + 1
        prev_p = display_pos_map.get(s["name"], i + 1)
        s["pos_change"] = prev_p - s["pos"]  # positivo = subió, negativo = bajó
        # Eliminar campos auxiliares de desempate
        s.pop("_excel_total", None)
        s.pop("_orig_idx", None)
        standings.append(s)

    # Posición provisional (incluye puntos de partidos en curso)
    # Desempate: si empatan en total_live, el que tiene mejor posición oficial gana
    live_order = sorted(standings, key=lambda x: (x["total_live"], -x["pos"]), reverse=True)
    for i, s in enumerate(live_order):
        s["live_pos"] = i + 1

    scoring_rules = _load_scoring_rules(ws1)

    # ── cuadro de honor (rows 249-258) ───────────────────────────────────────
    HONOR_ROWS = [
        {"row": 250, "title": "🥇 Campeón",           "category": "podium",   "short": "Campeón"},
        {"row": 251, "title": "🥈 Subcampeón",        "category": "podium",   "short": "Subcampeón"},
        {"row": 252, "title": "🥉 3er Puesto",        "category": "podium",   "short": "3er puesto"},
        {"row": 253, "title": "⚽ Bota de Oro",        "category": "scorers",  "short": "Bota de Oro"},
        {"row": 254, "title": "🥈 Bota de Plata",     "category": "scorers",  "short": "Bota de Plata"},
        {"row": 255, "title": "🥉 Bota de Bronce",    "category": "scorers",  "short": "Bota de Bronce"},
        {"row": 256, "title": "🏆 Balón de Oro",      "category": "players",  "short": "Balón de Oro"},
        {"row": 257, "title": "🥈 Balón de Plata",    "category": "players",  "short": "Balón de Plata"},
        {"row": 258, "title": "🥉 Balón de Bronce",   "category": "players",  "short": "Balón de Bronce"},
    ]
    honor_pts_rules = []
    for sec in scoring_rules.get("sections", []):
        if sec.get("key") == "honor":
            honor_pts_rules = [float(i["pts"]) for i in sec.get("items", [])]
            break

    honor = []
    honor_correct = {p["name"]: 0 for p in all_players}
    honor_filled  = {p["name"]: 0 for p in all_players}
    resolved_count = 0

    for idx, meta in enumerate(HONOR_ROWS):
        row = meta["row"]
        title = meta["title"]
        max_pts = honor_pts_rules[idx] if idx < len(honor_pts_rules) else None
        result_raw = _val(ws1, row, 13)
        actual = _parse_honor_actual(result_raw)
        if actual:
            resolved_count += 1

        preds = {}
        preds_list = []
        for p, ws in zip(all_players, all_ws):
            # Rows 250-252 (Podium) are dynamically computed in the KO bracket, so they are in _ko_preds
            # Rows 253-258 (Awards) were manually filled in the Group stage Excel and left blank in the KO stage
            if row in (250, 251, 252):
                pv_dict = _ko_preds.get(p["name"], {}).get("_honor", {})
                pv = pv_dict.get(str(row))
            else:
                pv = _val(ws, row, p["pred_col"])
            pred_raw = str(pv).strip() if pv and not str(pv).startswith("Pegar") else None
            pred = _normalize_honor_name(pred_raw) if pred_raw else None
            correct = bool(actual and pred and pred == actual)
            score = float(max_pts) if correct and max_pts is not None else 0.0
            if pred:
                honor_filled[p["name"]] += 1
            if correct:
                honor_correct[p["name"]] += 1
            entry = {
                "name": p["name"],
                "pred": pred,
                "score": score,
                "correct": correct,
                "color": PLAYER_COLORS[all_players.index(p) % len(PLAYER_COLORS)],
            }
            preds[p["name"]] = {"pred": pred, "score": score, "correct": correct}
            if pred:
                preds_list.append(entry)

        preds_list.sort(key=lambda x: (-x["score"], -int(x["correct"]), x["name"]))

        # Predicción más popular (consenso del grupo)
        pick_counts = {}
        for e in preds_list:
            pick_counts[e["pred"]] = pick_counts.get(e["pred"], 0) + 1
        consensus = max(pick_counts.items(), key=lambda x: x[1])[0] if pick_counts else None
        consensus_n = pick_counts.get(consensus, 0) if consensus else 0

        honor.append({
            "title": title,
            "short": meta["short"],
            "category": meta["category"],
            "max_pts": max_pts,
            "actual": actual,
            "resolved": actual is not None,
            "predictions": preds,
            "predictions_list": preds_list,
            "consensus": consensus,
            "consensus_count": consensus_n,
            "filled_count": len(preds_list),
        })

    honor_summary = {
        "total_items": len(HONOR_ROWS),
        "resolved": resolved_count,
        "pending": len(HONOR_ROWS) - resolved_count,
        "max_total_pts": float(_val(ws1, 62, 4) or 0),
        "by_player": sorted([
            {
                "name": p["name"],
                "honor_pts": float(all_clas.get(p["name"], {}).get("honor", 0) or 0),
                "correct": honor_correct[p["name"]],
                "filled": honor_filled[p["name"]],
                "color": PLAYER_COLORS[i % len(PLAYER_COLORS)],
            }
            for i, p in enumerate(all_players)
        ], key=lambda x: (-x["honor_pts"], -x["correct"], x["name"])),
    }

    # ── max points reference ─────────────────────────────────────────────────
    max_points = {
        "groups":    _val(ws1, 56, 4),
        "positions": _val(ws1, 57, 4),
        "q16":       None,
        "r16":       _val(ws1, 57, 4),
        "r8":        _val(ws1, 58, 4),
        "r4":        _val(ws1, 59, 4),
        "r2":        _val(ws1, 60, 4),
        "r34_final": _val(ws1, 61, 4),
        "honor":     _val(ws1, 62, 4),
    }

    weeks = _week_ranges_from_dates(spain_dates)
    prog_grid_ctx = {
        "player_q16_pts": player_q16_pts,
        "grid_preds": _build_player_grid_preds(all_players, all_ws),
        "pts": {
            "r8":    pts_r8_team,
            "r4":    pts_r4_team,
            "r2":    pts_r2_team,
            "r34":   pts_r34_team,
            "final": pts_final_team,
        },
        "actual": {
            "r8":    actual_r8_qualifiers,
            "r4":    actual_r4_qualifiers,
            "r2":    actual_r2_qualifiers,
            "r34":   actual_r34_qualifiers,
            "final": actual_final_qualifiers,
        },
    }
    progression = _build_daily_progression(
        matches, player_names, player_positions_pts, all_groups_finished,
        grid_ctx=prog_grid_ctx,
    )
    player_strengths = _build_player_strengths(matches, standings, player_names)

    from update_schedule import build_update_meta

    return {
        "meta": {
            "title":     "Porra Mundial 'Los Nanos' 2026",
            "generated": datetime.now().strftime("%d/%m/%Y %H:%M"),
            "update":    build_update_meta(),
            "players":   player_names,
            "live": {
                "active":  bool(live_match_names),
                "count":   len(live_match_names),
                "matches": live_match_names,
            },
            "colors":    {p["name"]: PLAYER_COLORS[i] for i, p in enumerate(all_players)},
            "weeks":     weeks,
            "scoring": {
                "sign":  pts_sign,
                "diff":  pts_diff,
                "exact": pts_exact,
            },
            "prizes": {"first": 40, "second": 20, "currency": "€"},
            "group_progress": {grp: count for grp, count in group_matches_count.items()},
        },
        "standings":   standings,
        "matches":     matches,
        "progression": progression,
        "honor":       honor,
        "honor_summary": honor_summary,
        "max_points":  max_points,
        "scoring_rules": scoring_rules,
        "player_strengths": player_strengths,
    }


def get_data():
    global _cache
    now = time.time()
    if _cache["data"] is None or (now - _cache["ts"]) > CACHE_TTL:
        try:
            data = build_data()
            # Inject highlights (data/highlights.json) into each match
            _hl_path = os.path.join(BASE, "data", "highlights.json")
            if os.path.isfile(_hl_path):
                try:
                    with open(_hl_path, encoding="utf-8") as _f:
                        _hl = json.load(_f)
                    for _m in data.get("matches", []):
                        _vid = _hl.get(_m.get("name", ""))
                        if _vid:
                            _m["highlights_video_id"] = _vid
                except (OSError, ValueError):
                    pass
            _cache["data"] = data
            _cache["error"] = None
        except Exception as e:
            _cache["error"] = str(e)
            if _cache["data"] is None:
                raise
        _cache["ts"] = now
    return _cache["data"]


# ── routes ───────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return send_from_directory(BASE, "index.html")


@app.route("/favicon.ico")
def favicon():
    return send_from_directory(os.path.join(BASE, "static"), "favicon.ico")


@app.route("/static/<path:filename>")
def static_files(filename):
    return send_from_directory(os.path.join(BASE, "static"), filename)


def _no_cache(response):
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response


@app.route("/api/data")
def api_data():
    try:
        return _no_cache(jsonify(get_data()))
    except Exception as e:
        import traceback
        traceback.print_exc()
        return _no_cache(jsonify({"error": str(e), "detail": "No se pudieron leer los Excel. Cierra Excel si los tienes abiertos e inténtalo de nuevo."})), 500


@app.route("/api/refresh")
def api_refresh():
    """Invalida la caché local para forzar una relectura del Excel.
    Solo afecta al servidor Flask local; no llama a la API externa."""
    _cache["ts"] = 0
    return _no_cache(jsonify({"ok": True, "ts": datetime.now().isoformat()}))


# ── Live match data proxy (scorers, results) ─────────────────────────────────
_wc_games_cache: dict = {"data": None, "ts": 0.0}
WC_GAMES_TTL = 300  # 5 minutes


@app.route("/api/wc_games")
def api_wc_games():
    global _wc_games_cache
    now = time.time()
    if _wc_games_cache["data"] is None or (now - _wc_games_cache["ts"]) > WC_GAMES_TTL:
        try:
            req = _urllib_req.Request(
                "https://worldcup26.ir/get/games",
                headers={"User-Agent": "PorraNanos/1.0"},
            )
            with _urllib_req.urlopen(req, timeout=12) as r:
                raw = json.load(r)
            games = raw.get("games", raw) if isinstance(raw, dict) else raw
            _wc_games_cache["data"] = games
            _wc_games_cache["ts"] = now
        except Exception as exc:
            if _wc_games_cache["data"] is not None:
                pass  # serve stale cache on network error
            else:
                return _no_cache(jsonify({"error": str(exc)})), 503
    return _no_cache(jsonify(_wc_games_cache["data"]))


if __name__ == "__main__":
    print("\n🏆  Porra Mundial 'Los Nanos' 2026")
    print("   http://localhost:5050\n")
    app.run(host="0.0.0.0", port=5050, debug=False)
